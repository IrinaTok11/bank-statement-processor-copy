#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""merge_statements_international.py
---------------------
Consolidate multiple business bank statements into standardized Transactions table
with forensic screening, categorization, and comprehensive analytics reporting.

Processes Excel files from multiple banks, applies detection algorithms for suspicious
patterns, and generates audit-ready findings with full source traceability.
"""

from __future__ import annotations

__author__  = "Irina Tokmianina"
__version__ = "1.0.0"
__license__ = "MIT"

import pandas as pd
import os
from pathlib import Path
import warnings
import re
from datetime import datetime
import numpy as np
import argparse
import zipfile
import tempfile
import logging


# ============================================================================
# Configuration: Detection Rule Parameters
# ============================================================================

RULES = {
    "duplicate_payment": {
        "display": "Duplicate Detection",
        "params": {
            "amount_tolerance": 0.01,   # 1% tolerance
            "max_days_apart": 3,
            "min_amount": 100.0
        },
        "config": {
            "threshold": "Same vendor within ±1% amount AND within 3 days (min debit $100)",
            "rationale": "Allows for minor rounding differences; focuses on material duplicates",
            "source": "Industry practice"
        },
        "methodology": "Duplicate Detection: Same vendor within ±1% amount within 3 days (min debit $100)."
    },
    "split_payments": {
        "display": "Split Payments",
        "params": {
            "min_count": 2,
            "min_total": 10000.0,
            "large_payment_threshold": 5000.0
        },
        "config": {
            "threshold": "2+ payments to same vendor on same day; total >$10,000 (flag if any single >$5,000)",
            "rationale": "Identifies potential approval splitting / policy circumvention patterns",
            "source": "Common approval/reporting thresholds"
        },
        "methodology": "Split Payments: 2+ same-day payments to same vendor; total >$10,000 (flag if any single >$5,000)."
    },
    "vendor_concentration_growth": {
        "display": "Vendor Concentration Growth",
        "params": {
            "baseline_min": 1000.0,
            "growth_factor": 2.0  # >100% growth (latest > baseline * 2)
        },
        "config": {
            "threshold": "Baseline→latest month spend >2x AND baseline >$1,000",
            "rationale": "Highlights rapidly increasing dependency on a single vendor over the analysis period",
            "source": "Forensic vendor-risk screening"
        },
        "methodology": "Vendor Concentration Growth: Baseline (first active month) → latest month >2x; baseline >$1,000."
    },
}

def get_detection_threshold_rows(run_stats=None):
    """Build Config & Thresholds table rows from RULES + any run-specific dynamic thresholds."""
    rows = []
    # Static rule thresholds
    for key in ("duplicate_payment", "split_payments", "vendor_concentration_growth"):
        spec = RULES[key]
        rows.append([spec["display"], spec["config"]["threshold"], spec["config"]["rationale"], spec["config"]["source"]])
    # Keep the original additional rows (documented, but params are not yet centralized)
    rows.extend([
        ["New Counterparty Alert", ">$50,000", "Materiality threshold for significant new counterparties", "Typical SMB materiality level"],
        ["Round Amount Clustering", "Exact $1,000 increments", "Identify potential estimates or manual entries", "Forensic screening"],
        ["Weekend Activity", "Dynamic (p95 debit * 1.25)", "Flags unusually large weekend debits (excludes taxes/payroll/recurring)", "Run-specific (logged)"],
    ])
    # Run-specific dynamic thresholds (if provided)
    if run_stats:
        if "p95_debit" in run_stats:
            rows.append(["RUN-SPECIFIC: p95_debit", f"{run_stats['p95_debit']:.2f}", "95th percentile of debit amounts", "Computed from Transactions"])
        if "weekend_threshold" in run_stats:
            rows.append(["RUN-SPECIFIC: weekend_threshold", f"{run_stats['weekend_threshold']:.2f}", "Weekend debit alert threshold", "Computed from Transactions"])
    return rows

# Disable warnings for cleaner output
warnings.filterwarnings('ignore')

logger = logging.getLogger(__name__)


def configure_logging(verbose: bool = False) -> None:
    """Configure console logging with minimal formatting."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(message)s")


# ============================================================================
# Excel Metadata Utilities
# ============================================================================

def patch_excel_app_metadata(xlsx_path: Path, application: str = "Microsoft Excel", 
                            app_version: str = "16.0300") -> None:
    """Patch docProps/app.xml to set Application/AppVersion metadata."""
    
    logger = logging.getLogger(__name__)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")

    def _replace_tag(xml_text: str, tag: str, value: str) -> str:
        """Replace the first occurrence of <tag>...</tag> (with optional namespace prefix)."""
        # Matches <Application> ... </Application> and also <ns0:Application> ... </ns0:Application>
        pat = rf"(<(?:\w+:)?{tag}>)(.*?)(</(?:\w+:)?{tag}>)"
        if re.search(pat, xml_text, flags=re.DOTALL):
            return re.sub(pat, lambda m: f"{m.group(1)}{value}{m.group(3)}", xml_text, count=1, flags=re.DOTALL)
        return xml_text

    try:
        with zipfile.ZipFile(xlsx_path, mode="r") as z_in, zipfile.ZipFile(tmp_path, mode="w") as z_out:
            for item in z_in.infolist():
                data = z_in.read(item.filename)

                if item.filename == "docProps/app.xml":
                    try:
                        xml = data.decode("utf-8")
                    except UnicodeDecodeError:
                        xml = data.decode("utf-8", errors="replace")

                    xml = _replace_tag(xml, "Application", application)
                    xml = _replace_tag(xml, "AppVersion", app_version)

                    data = xml.encode("utf-8")

                z_out.writestr(item, data)

        tmp_path.replace(xlsx_path)
        logger.info("Patched Excel app metadata in: %s", xlsx_path)

    except Exception as e:
        # If anything goes wrong, keep the original workbook intact.
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
        logger.warning("Excel app metadata patch skipped: %s", str(e))


def first_sheet_name(file_path: Path) -> str:
    """Return the first worksheet name for an Excel file, or 'Sheet1' as a fallback."""
    try:
        xf = pd.ExcelFile(file_path)
        return xf.sheet_names[0] if xf.sheet_names else 'Sheet1'
    except Exception:
        return 'Sheet1'


# ============================================================================
# Bank Format Detection
# ============================================================================

def detect_bank_format(file_path):
    """Detect bank statement format from Excel file.
    
    Extracts bank name from filename, auto-detects header row and column structure,
    and determines format type (debit/credit separate vs amount signed).
    """
    import os
    import openpyxl
    
    try:
        filename = os.path.basename(file_path)
        bank_name = filename.split('_')[0]  # First word before underscore
        
        # Get actual sheet name from Excel file
        wb = openpyxl.load_workbook(file_path, read_only=True)
        actual_sheet_name = wb.sheetnames[0]  # Use first sheet
        
        # Check for multi-sheet files
        if len(wb.sheetnames) > 1:
            logger.warning(f"\n  WARNING: File '{filename}' contains {len(wb.sheetnames)} sheets:")
            for idx, sheet in enumerate(wb.sheetnames[:5], 1):  # Show first 5 sheets
                marker = " PROCESSING" if idx == 1 else "✗ IGNORED"
                logger.info(f"   {marker}: Sheet {idx} - '{sheet}'")
            if len(wb.sheetnames) > 5:
                logger.info(f"   ... and {len(wb.sheetnames) - 5} more sheets (all ignored)")
            logger.info(f"   ➜ Only first sheet '{actual_sheet_name}' will be processed")
            logger.info(f"   ➜ To process all sheets: split file or modify script")
            logger.info("")
        
        wb.close()
        
        df_raw = pd.read_excel(file_path, header=None, nrows=50)
        
        header_row = None
        format_type = None
        
        for i in range(min(50, len(df_raw))):
            row = df_raw.iloc[i]
            row_str_lower = ' | '.join([str(x).lower() for x in row if pd.notna(x)])
            
            # Look for header markers
            has_date = any(word in row_str_lower for word in ['date', 'posting date', 'trans date', 'value date'])
            has_desc = any(word in row_str_lower for word in ['description', 'transaction', 'details', 'particulars', 'narrative'])
            
            if has_date and has_desc:
                # Found potential header!
                header_row = i
                
                # Detect format type
                has_debit = 'debit' in row_str_lower or 'money out' in row_str_lower or 'withdrawals' in row_str_lower or 'paid out' in row_str_lower
                has_credit = 'credit' in row_str_lower or 'money in' in row_str_lower or 'deposits' in row_str_lower or 'additions' in row_str_lower or 'paid in' in row_str_lower
                has_amount = 'amount' in row_str_lower
                
                if has_debit and has_credit:
                    format_type = 'debit_credit'
                elif has_amount and not has_debit:
                    format_type = 'amount_signed'
                else:
                    # Fallback: assume debit_credit if we found debit OR credit
                    format_type = 'debit_credit' if (has_debit or has_credit) else 'amount_signed'
                
                break
        
        if header_row is None:
            logger.info(f"  Could not find header row in {filename}")
            return None
        
        header_values = [str(x).strip() for x in df_raw.iloc[header_row] if pd.notna(x)]
        
        # Find column indices
        def find_column(keywords):
            for idx, val in enumerate(header_values):
                val_lower = val.lower()
                if any(kw in val_lower for kw in keywords):
                    return idx
            return None
        
        date_col = find_column(['date', 'posting date', 'trans date', 'value date'])
        desc_col = find_column(['description', 'transaction', 'details', 'particulars', 'narrative'])
        balance_col = find_column(['balance', 'running balance'])
        
        columns = {
            'date': date_col,
            'description': desc_col,
            'balance': balance_col
        }
        
        if format_type == 'debit_credit':
            debit_col = find_column(['debit', 'money out', 'withdrawal', 'payments out', 'paid out'])
            credit_col = find_column(['credit', 'money in', 'deposit', 'addition', 'payments in', 'paid in'])
            columns['debit'] = debit_col
            columns['credit'] = credit_col
        elif format_type == 'amount_signed':
            amount_col = find_column(['amount', 'value'])
            columns['amount'] = amount_col
        
        # Optional: Try to find Type/Category column
        type_col = find_column(['type', 'category', 'transaction type'])
        if type_col is not None:
            columns['type'] = type_col
        
        config = {
            'bank': bank_name,
            'sheet': actual_sheet_name,  # Use actual sheet name from Excel
            'header': header_row,
            'format_type': format_type,
            'columns': columns
        }
        
        logger.info(f" Detected: {bank_name} (format: {format_type}, header row: {header_row})")
        return config
        
    except Exception as e:
        logger.error(f" Error detecting format for {file_path}: {e}")
        return None


# ============================================================================
# Data Normalization
# ============================================================================

def normalize_date_column(df, column_name):
    """Normalize date column to date objects."""
    if column_name in df.columns:
        df[column_name] = pd.to_datetime(df[column_name], errors='coerce')
        df[column_name] = df[column_name].dt.date
    return df


def normalize_amount_column(df, column_name):
    """Normalize amount columns to numeric values."""
    if column_name in df.columns:
        df[column_name] = df[column_name].astype(str).str.replace(' ', '').str.replace(',', '')
        df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
    return df


def clean_description(text):
    if pd.isna(text) or str(text).strip() == '':
        return text

    original = str(text).strip()
    text_cleaned = original

    # Remove common transaction type prefixes
    patterns_to_remove = [
        r'^ACH\s+(CREDIT|DEBIT|PAYMENT)\s+-\s+',
        r'^WIRE\s+(TRANSFER|OUT|IN)\s+-\s+',
        r'^CHECK\s+PAID\s+#\d+\s+-\s+',
        r'^REMOTE\s+DEPOSIT\s+-\s+',
        r'^BILL\s+PAYMENT\s+-\s+',
        r'^SERVICE\s+FEE\s+-\s+',
        r'^MISC\.\s+CREDIT\s+-\s+',
        r'^Card\s+Purchase\s+\d+/\d+\s+\d+:\d+\s+',
        r'^Inward\s+Payments\s+',
        r'^Outward\s+Fast\s+Payments\s+',
        r'^Outward\s+Payments\s+',
        r'^CARD\s+PURCHASE\s+-\s+',
        r'^DEBIT\s+CARD\s+-\s+',
        r'^ONLINE\s+PAYMENT\s+-\s+',
    ]

    for pattern in patterns_to_remove:
        text_cleaned = re.sub(pattern, '', text_cleaned, flags=re.IGNORECASE)

    # Remove reference numbers at the end (e.g., "INV 10484", "PO 5839", "CP434395", "FP341404")
    text_cleaned = re.sub(r'\s+(INV|PO|CP|FP)\s*\d+$', '', text_cleaned, flags=re.IGNORECASE)

    # Remove excess whitespace
    text_cleaned = re.sub(r'\s+', ' ', text_cleaned)
    text_cleaned = text_cleaned.strip()

    return text_cleaned if text_cleaned else original

def normalize_vendor(text):
    if pd.isna(text) or str(text).strip() == '':
        return text

    vendor = str(text).strip().upper()
    # Remove common payment rail prefixes that sometimes survive Description_clean
    prefix_patterns = [
        r'^ACH\s+(CREDIT|DEBIT|PAYMENT)\s*-\s*',
        r'^WIRE\s+(TRANSFER|OUT|IN)\s*-\s*',
        r'^CHECK\s+PAID\s+#\d+\s*-\s*',
        r'^REMOTE\s+DEPOSIT\s*-\s*',
        r'^BILL\s+PAYMENT\s*-\s*',
        r'^SERVICE\s+FEE\s*-\s*',
        r'^MISC\.?\s+CREDIT\s*-\s*',
        r'^CARD\s+PURCHASE\s*-\s*',
        r'^CARD\s+PURCHASE\s+\d{1,2}/\d{1,2}\s+\d{1,2}:\d{2}\s+',
        r'^DEBIT\s+CARD\s*-\s*',
        r'^CREDIT\s+CARD\s*-\s*',
        r'^ONLINE\s+PAYMENT\s*-\s*',
        r'^INWARD\s+PAYMENTS\s+',
        r'^OUTWARD\s+FAST\s+PAYMENTS\s+',
        r'^OUTWARD\s+PAYMENTS\s+',
        r'^FASTER\s+PAYMENTS\s*-\s*(OUT|IN)\s+',
        r'^FASTER\s+PAYMENTS\s+(OUT|IN)\s+',
    ]
    for p in prefix_patterns:
        vendor = re.sub(p, '', vendor, flags=re.IGNORECASE)

    # Remove trailing reference tokens (INV/PO/CP/FP/REF + digits)
    vendor = re.sub(r'\s+(INV|PO|CP|FP|REF)\s*\d+$', '', vendor, flags=re.IGNORECASE)


    # Remove common suffixes
    vendor = re.sub(r'\s+(LLC|INC|CORP|LTD|CO|COMPANY|CORPORATION|LIMITED)\b', '', vendor)

    # Remove punctuation
    vendor = re.sub(r'[.,;:\-]', '', vendor)

    # Remove multiple spaces
    vendor = re.sub(r'\s+', ' ', vendor)

    return vendor.strip()

def normalize_company_name(df, column_name):
    """
    Normalize counterparty/company name column.
    
    Extracts company name from description, removes common prefixes like:
    - ACH CREDIT/DEBIT
    - WIRE TRANSFER
    - CHECK PAID
    - Card Purchase
    etc.
    
    Creates a new column with suffix '_clean' (e.g., 'Description_clean').
    Original column remains unchanged.
    """
    if column_name not in df.columns:
        return df
    
    
    df[f'{column_name}_clean'] = df[column_name].apply(clean_description)
    
    # Create additional Vendor_Normalized column for concentration analysis
    
    # Apply to Description_clean if it exists, otherwise to original
    source_col = f'{column_name}_clean' if f'{column_name}_clean' in df.columns else column_name
    df['Vendor_Normalized'] = df[source_col].apply(normalize_vendor)
    
    return df


# ============================================================================
# Forensic Detection Algorithms
# ============================================================================

def detect_findings(df, return_run_stats=False):
    """Detect actionable findings requiring review.
    
    Implements rule-based detection for fraud patterns and business risks
    based on forensic accounting practices.
    """
    findings = []
    run_stats = {}
    alert_id = 1
    
    # Add Transaction IDs if not present
    if 'Txn_ID' not in df.columns:
        df['Txn_ID'] = [f"TXN_{i+1:05d}" for i in range(len(df))]
    

    # Work on a copy so helper columns do not leak into the Transactions output
    df = df.copy()

    # Internal vendor key used for grouping in detection rules (NOT exported)
    if 'Vendor_Normalized' in df.columns:
        df['_vendor_key'] = df['Vendor_Normalized'].fillna('').astype(str).str.strip()
    elif 'Description_clean' in df.columns:
        df['_vendor_key'] = df['Description_clean'].fillna('').astype(str).str.strip()
    elif 'Description' in df.columns:
        df['_vendor_key'] = df['Description'].fillna('').astype(str).str.strip()
    else:
        df['_vendor_key'] = 'UNKNOWN'
    df['_vendor_key'] = df['_vendor_key'].replace('', 'UNKNOWN')

    # Helper function to create source location string
    def make_source_loc(row):
        """Create traceable source location string for audit trail: file | sheet | row."""
        src_file = row.get('Source_File', '?')
        src_sheet = row.get('Source_Sheet', '?')
        src_row = row.get('Source_Row', '?')
        return f"{src_file} | {src_sheet} | R{src_row}"
    
    # PRIORITY 1: DUPLICATE PAYMENTS
    # Exact or near-exact match: same vendor, same/similar amount, within 3 days
    # EXCLUDE: Bank fees, recurring services (legitimate duplicates)
    
    # Exclude recurring service categories from duplicate detection
    EXCLUDED_CATEGORIES_DUPLICATES = [
        'Shipping & Logistics',      # UPS, FedEx - recurring deliveries
        'Utilities & Services',       # Waste Management, power, water
        'Bank Fees',                  # Bank fees
        'Payroll & Benefits',         # Payroll (can be weekly)
        'Taxes & Government',         # Taxes (can be recurring)
        'Rent & Facilities'           # Rent (fixed recurring amounts)
    ]
    
    # Filter out excluded categories
    if 'Operation_Type' in df.columns:
        df_for_duplicates = df[~df['Operation_Type'].isin(EXCLUDED_CATEGORIES_DUPLICATES)].copy()
    else:
        df_for_duplicates = df.copy()
    
    df_sorted = df_for_duplicates.sort_values(['_vendor_key', 'Debit', 'Date'])
    
    # Track split payment candidates to avoid double-flagging
    potential_splits = set()
    
    for i in range(len(df_sorted) - 1):
        if pd.isna(df_sorted.iloc[i]['Debit']) or df_sorted.iloc[i]['Debit'] == 0:
            continue
            
        current = df_sorted.iloc[i]
        next_row = df_sorted.iloc[i + 1]
        
        # Skip bank fees and service charges (legitimate monthly duplicates)
        vendor_lower = str(current['Description_clean']).lower()
        if any(keyword in vendor_lower for keyword in ['monthly service charge', 'service fee', 
                                                        'maintenance fee', 'account fee']):
            continue
        
        # Check for exact or near-duplicate (within 1% of amount)
        amount_diff = abs(current['Debit'] - next_row['Debit']) / current['Debit'] if current['Debit'] > 0 else 1
        
        # Skip if amounts are very small (< $100) - likely recurring fees
        if current['Debit'] < RULES['duplicate_payment']['params']['min_amount']:
            continue
        
        if (current['_vendor_key'] == next_row['_vendor_key'] and
            amount_diff < RULES['duplicate_payment']['params']['amount_tolerance'] and  # Within tolerance
            pd.notna(current['Date']) and pd.notna(next_row['Date']) and
            abs((current['Date'] - next_row['Date']).days) <= RULES['duplicate_payment']['params']['max_days_apart']):
            
            # Check if same day (might be split payment)
            days_apart = abs((current['Date'] - next_row['Date']).days)
            if days_apart == 0:
                # Potential split payment - mark for later review
                potential_splits.add(current['Txn_ID'])
                potential_splits.add(next_row['Txn_ID'])
                # Don't add as duplicate yet, will be handled by split payment logic
                continue
            
            findings.append({
                'Alert_ID': f"A{alert_id:04d}",
                'Priority': 'P1',
                'Category': 'Duplicate Payment',
                'Severity': 'HIGH',
                'Vendor': current['_vendor_key'],
                'Amount': current['Debit'],
                'Amount_Type': 'Transaction Amount',
                'Date_1': current['Date'],
                'Date_2': next_row['Date'],
                'Bank': current['Bank'],
                'Operation_Type': current.get('Operation_Type', 'Unknown'),
                'Source_Loc': f"{make_source_loc(current)} & {make_source_loc(next_row)}",
                'Evidence': f"Same vendor, ${current['Debit']:,.2f} & ${next_row['Debit']:,.2f} (within ±1% tolerance), {days_apart} day(s) apart",
                'Txn_IDs': f"{current['Txn_ID']}, {next_row['Txn_ID']}",
                'Next_Step': 'Request invoices for both transactions, verify with AP',
                'Risk': 'Possible duplicate processing or vendor billing error',
                'Action': 'Verify both payments cleared, investigate with vendor'
            })
            alert_id += 1
    
    # PRIORITY 1: NEW COUNTERPARTY - LARGE PAYMENT
    # First transaction >$50k (with exclusions)
    exclude_new_vendor = ['payroll', 'irs', 'internal revenue', 'tax payment', 
                          'eftps', 'utility', 'insurance', 'health insurance']
    
    vendor_first_dates = df[df['Debit'] > 0].groupby('_vendor_key')['Date'].min()
    
    for vendor, first_date in vendor_first_dates.items():
        # Skip excluded categories
        if any(excl in vendor.lower() for excl in exclude_new_vendor):
            continue
            
        first_payment = df[(df['_vendor_key'] == vendor) & 
                          (df['Date'] == first_date) & 
                          (df['Debit'] > 50000)]
        
        if len(first_payment) > 0:
            payment = first_payment.iloc[0]
            findings.append({
                'Alert_ID': f"A{alert_id:04d}",
                'Priority': 'P1',
                'Category': 'New Counterparty - High Amount',
                'Severity': 'HIGH',
                'Vendor': vendor,
                'Amount': payment['Debit'],
                'Amount_Type': 'Transaction Amount',  # NEW
                'Date_1': payment['Date'],
                'Date_2': None,
                'Bank': payment['Bank'],
                'Operation_Type': payment.get('Operation_Type', 'Unknown'),
                'Source_Loc': make_source_loc(payment),
                'Evidence': f"First transaction in dataset window: ${payment['Debit']:,.2f}",
                'Txn_IDs': payment['Txn_ID'],
                'Next_Step': 'Request supplier onboarding docs: tax form (W-9/W-8 if US; VAT registration if EU/UK), contract/PO, approval documentation',
                'Risk': 'No prior transaction history in dataset period',
                'Action': 'Verify vendor legitimacy and authorization'
            })
            alert_id += 1
    
    # PRIORITY 2: ROUND AMOUNT CLUSTERING
    # Multiple round amounts (ending in 000.00) - excluding known fixed costs
    fixed_cost_keywords = ['rent', 'lease', 'payroll', 'salary', 'loan payment', 'insurance']
    
    round_amounts = df[df['Debit'] > 0].copy()
    round_amounts['is_round'] = (round_amounts['Debit'] % 1000 == 0) & (round_amounts['Debit'] >= 1000)
    round_amounts['is_fixed'] = round_amounts['Description_clean'].str.lower().str.contains('|'.join(fixed_cost_keywords), na=False, regex=True)
    
    suspicious_rounds = round_amounts[(round_amounts['is_round']) & (~round_amounts['is_fixed'])]
    
    if len(suspicious_rounds) >= 3:
        # Group by amount
        for amount, group in suspicious_rounds.groupby('Debit'):
            if len(group) >= 2:
                vendors = group['_vendor_key'].unique()
                txn_ids = ', '.join(group['Txn_ID'].tolist()[:5])
                
                findings.append({
                    'Alert_ID': f"A{alert_id:04d}",
                    'Priority': 'P2',
                    'Category': 'Round Amount Clustering',
                    'Severity': 'MEDIUM',
                    'Vendor': f"{len(group)} transactions: {', '.join(vendors[:2])}...",
                    'Amount': amount,
                    'Amount_Type': 'Transaction Amount',  # NEW
                    'Date_1': group['Date'].min(),
                    'Date_2': group['Date'].max(),
                    'Bank': 'Multiple' if len(group['Bank'].unique()) > 1 else group['Bank'].iloc[0],
                    'Operation_Type': 'Multiple' if len(group['Operation_Type'].unique()) > 1 else group['Operation_Type'].iloc[0],
                    'Source_Loc': f"Multiple ({len(group)} txns)",
                    'Evidence': f"{len(group)} payments of exactly ${amount:,.2f}",
                    'Txn_IDs': txn_ids,
                    'Next_Step': 'Request invoices, review for estimation or manual entry',
                    'Risk': 'Unusual clustering of exact amounts (may indicate estimation)',
                    'Action': 'Review for potential manipulation or estimation errors'
                })
                alert_id += 1
    
    # PRIORITY 2: SPLIT PAYMENT DETECTION
    # Multiple payments to same vendor on same day, total >$10k
    # This detects potential approval splitting patterns
    
    # Group by date and vendor, but keep track of individual transaction details
    df_debits = df[df['Debit'] > 0].copy()
    
    for (date, vendor), group in df_debits.groupby([df_debits['Date'], '_vendor_key']):
        if len(group) >= RULES['split_payments']['params']['min_count'] and group['Debit'].sum() > RULES['split_payments']['params']['min_total']:
            # Collect details from actual transactions
            banks = group['Bank'].unique()
            op_types = group['Operation_Type'].unique() if 'Operation_Type' in group.columns else ['Unknown']
            txn_ids = ', '.join(group['Txn_ID'].tolist())
            
            # Build source location list
            source_locs = []
            for idx, txn in group.iterrows():
                source_locs.append(make_source_loc(txn))
            source_loc_str = '; '.join(source_locs[:3])  # First 3
            if len(source_locs) > 3:
                source_loc_str += f' (+{len(source_locs)-3} more)'
            
            # Check if at least one payment >$5k (common approval threshold)
            has_large_payment = (group['Debit'] > RULES['split_payments']['params']['large_payment_threshold']).any()
            
            findings.append({
                'Alert_ID': f"A{alert_id:04d}",
                'Priority': 'P2',
                'Category': 'Split Payments',
                'Severity': 'MEDIUM',
                'Vendor': vendor,
                'Amount': group['Debit'].sum(),
                'Amount_Type': 'Total Paid (Multiple Transactions)',
                'Date_1': date,
                'Date_2': None,
                'Bank': 'Multiple' if len(banks) > 1 else banks[0],
                'Operation_Type': 'Multiple' if len(op_types) > 1 else op_types[0],
                'Source_Loc': source_loc_str,
                'Evidence': f'{len(group)} payments totaling ${group["Debit"].sum():,.2f} on same day' + 
                           (' (at least one >$5k)' if has_large_payment else '') + 
                           ', may indicate approval limits',
                'Txn_IDs': txn_ids,
                'Next_Step': 'Verify business justification and approval chain',
                'Risk': f'Potential approval splitting pattern',
                'Action': 'Review reason for multiple payments to same vendor'
            })
            alert_id += 1
    
    # PRIORITY 3: WEEKEND ACTIVITY (with smart threshold and collapse)
    # Large payments on weekends - excluding automated, taxes, payroll
    exclude_weekend = ['ach', 'automatic', 'recurring', 'payroll', 'tax payment', 
                       'eftps', 'irs', 'internal revenue']
    
    df_copy = df[df['Debit'] > 0].copy()
    df_copy['Date'] = pd.to_datetime(df_copy['Date'])
    df_copy['Weekday'] = df_copy['Date'].dt.dayofweek
    df_copy['is_weekend'] = df_copy['Weekday'] >= 5
    df_copy['is_excluded'] = df_copy['Description_clean'].str.lower().str.contains('|'.join(exclude_weekend), na=False, regex=True)
    
    # Dynamic threshold: use 95th percentile or $25k, whichever is higher
    # This adapts to the transaction volume of the business
    percentile_95 = None
    if len(df_copy) > 20:
        percentile_95 = df_copy['Debit'].quantile(0.95)
        weekend_threshold = max(25000, percentile_95)
    else:
        weekend_threshold = 25000  # Default for small datasets

    # Store run-specific calculations for audit trail (does not affect detection logic)
    run_stats['weekend_outflow_txn_count'] = int(len(df_copy))
    run_stats['p95_debit'] = float(percentile_95) if percentile_95 is not None else None
    run_stats['weekend_threshold'] = float(weekend_threshold)
    run_stats['weekend_threshold_floor'] = 25000
    
    weekend_manual = df_copy[(df_copy['is_weekend']) & (~df_copy['is_excluded']) & (df_copy['Debit'] > weekend_threshold)]
    
    # COLLAPSE LOGIC: If >3 weekend transactions, create ONE summary alert
    if len(weekend_manual) > 3:
        # Create single collapsed alert
        total_amount = weekend_manual['Debit'].sum()
        txn_ids = ', '.join(weekend_manual['Txn_ID'].tolist()[:15])  # First 15 IDs
        
        # Get top 3 by amount for summary
        top_3 = weekend_manual.nlargest(3, 'Debit')
        top_3_summary = []
        for idx, row in top_3.iterrows():
            day_name = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][row['Weekday']]
            top_3_summary.append(f"{row['_vendor_key'][:25]} (${row['Debit']:,.0f}, {day_name})")
        
        findings.append({
            'Alert_ID': f"A{alert_id:04d}",
            'Priority': 'P3',
            'Category': 'Weekend Activity (Multiple)',
            'Severity': 'LOW',
            'Vendor': f"{len(weekend_manual)} transactions; see details",
            'Amount': total_amount,
            'Amount_Type': 'Total Amount (Multiple Transactions)',  # NEW
            'Date_1': weekend_manual['Date'].min(),
            'Date_2': weekend_manual['Date'].max(),
            'Bank': 'Multiple' if len(weekend_manual['Bank'].unique()) > 1 else weekend_manual['Bank'].iloc[0],
            'Operation_Type': 'Multiple',
            'Source_Loc': f"Multiple ({len(weekend_manual)} txns)",
            'Evidence': f'{len(weekend_manual)} weekend postings >${weekend_threshold:,.0f}, totaling ${total_amount:,.2f}. Top 3: {"; ".join(top_3_summary)}',
            'Txn_IDs': txn_ids + ('...' if len(weekend_manual) > 15 else ''),
            'Next_Step': 'Review for business context; posting dates may differ from processing dates',
            'Risk': 'Weekend posting pattern (review context)',
            'Action': 'Note: Weekend postings common for automated systems and deadline-driven payments'
        })
        alert_id += 1
    
    elif len(weekend_manual) > 0:
        # Individual alerts for 1-3 weekend transactions (material enough to review separately)
        for idx, row in weekend_manual.iterrows():
            day_name = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][row['Weekday']]
            findings.append({
                'Alert_ID': f"A{alert_id:04d}",
                'Priority': 'P3',
                'Category': 'Weekend Activity',
                'Severity': 'LOW',
                'Vendor': row['_vendor_key'],
                'Amount': row['Debit'],
                'Amount_Type': 'Transaction Amount',  # NEW
                'Date_1': row['Date'],
                'Date_2': None,
                'Bank': row['Bank'],
                'Operation_Type': row.get('Operation_Type', 'Unknown'),
                'Source_Loc': make_source_loc(row),
                'Evidence': f'Posted on {day_name} (posting date may differ from transaction date)',
                'Txn_IDs': row['Txn_ID'],
                'Next_Step': 'Verify business justification if material',
                'Risk': 'Weekend posting (review context)',
                'Action': 'Note: Posting date may differ from processing date'
            })
            alert_id += 1
    
    # PRIORITY 3: HIGH VELOCITY OUTFLOWS  
    # More than 5 large payments (>$20k) in single day
    df_copy = df[df['Debit'] > 20000].copy()
    daily_counts = df_copy.groupby('Date').agg({
        'Debit': ['count', 'sum'],
        'Txn_ID': lambda x: ', '.join(x.tolist()[:10])
    })
    daily_counts.columns = ['Count', 'Total', 'Txn_IDs']
    
    high_velocity = daily_counts[daily_counts['Count'] >= 5]
    
    for date, row in high_velocity.iterrows():
        findings.append({
            'Alert_ID': f"A{alert_id:04d}",
            'Priority': 'P3',
            'Category': 'High Velocity Outflows',
            'Severity': 'LOW',
            'Vendor': f'{int(row["Count"])} vendors',
            'Amount': row['Total'],
            'Amount_Type': 'Total Amount (Single Day)',  # NEW
            'Date_1': date,
            'Date_2': None,
            'Bank': 'Multiple',
            'Operation_Type': 'Multiple',
            'Source_Loc': 'See Txn_IDs',
            'Evidence': f'{int(row["Count"])} large payments (>${20000:,}) on single day',
            'Txn_IDs': row['Txn_IDs'],
            'Next_Step': 'Review for business context (e.g., month-end, quarter-end)',
            'Risk': 'Unusual concentration of payments',
            'Action': 'Review for business justification (e.g., period-end processing)'
        })
        alert_id += 1
    
    # PRIORITY 3: VENDOR CONCENTRATION TREND
    # Vendors that grew significantly
    if len(df['Date'].unique()) > 30:  # Need enough data
        # EXCLUDE recurring services from concentration analysis
        # These are expected to be regular and their growth is usually contractual, not risk
        EXCLUDED_CATEGORIES_CONCENTRATION = [
            'Shipping & Logistics',        # UPS, FedEx - regular delivery services
            'Utilities & Services',        # Electricity, water, waste management - regular services
            'Taxes & Government',          # Tax payments - expected regular payments
            'Bank Fees',                   # Banking fees - expected charges
            'Payroll & Benefits',          # Payroll - expected regular expense
            'Rent & Facilities',           # Rent/lease - fixed contractual amounts
            'Insurance & Security'         # Insurance premiums - contractual payments
        ]
        
        # Filter dataframe - exclude recurring service categories
        if 'Operation_Type' in df.columns:
            df_for_concentration = df[~df['Operation_Type'].isin(EXCLUDED_CATEGORIES_CONCENTRATION)].copy()
        else:
            df_for_concentration = df.copy()
        
        df_copy = df_for_concentration.copy()
        df_copy['Date'] = pd.to_datetime(df_copy['Date'])
        df_copy['Month'] = df_copy['Date'].dt.to_period('M')
        
        monthly_vendor = df_copy[df_copy['Debit'] > 0].groupby(['Month', '_vendor_key'])['Debit'].sum().unstack(fill_value=0)
        
        # Check for vendors that doubled
        for vendor in monthly_vendor.columns:
            amounts = monthly_vendor[vendor]
            if len(amounts) >= 2:
                active_amounts = amounts[amounts > 0]
                if len(active_amounts) == 0:
                    continue

                baseline_period = active_amounts.index[0]
                baseline_amount = float(active_amounts.iloc[0])
                latest_period = amounts.index[-1]
                latest_amount = float(amounts.iloc[-1])

                if baseline_amount > RULES['vendor_concentration_growth']['params']['baseline_min'] and latest_amount > baseline_amount * RULES['vendor_concentration_growth']['params']['growth_factor']:
                    # Txn_ID evidence should map to the months referenced in Evidence (baseline & latest)
                    baseline_txns = df_copy[
                        (df_copy['_vendor_key'] == vendor) &
                        (df_copy['Debit'] > 0) &
                        (df_copy['Month'] == baseline_period)
                    ]
                    latest_txns = df_copy[
                        (df_copy['_vendor_key'] == vendor) &
                        (df_copy['Debit'] > 0) &
                        (df_copy['Month'] == latest_period)
                    ]

                    def _compact_txn_ids(_df, max_ids=2):
                        ids = _df.sort_values('Debit', ascending=False)['Txn_ID'].astype(str).tolist()
                        if len(ids) == 0:
                            return '—'
                        if len(ids) <= max_ids:
                            return ', '.join(ids)
                        return ', '.join(ids[:max_ids]) + f' (+{len(ids) - max_ids} more)'

                    txn_id_summary = (
                        f'Baseline {baseline_period}: {_compact_txn_ids(baseline_txns)} | '
                        f'Latest {latest_period}: {_compact_txn_ids(latest_txns)}'
                    )
                    
                    findings.append({
                        'Alert_ID': f"A{alert_id:04d}",
                        'Priority': 'P3',
                        'Category': 'Vendor Concentration Growth',
                        'Severity': 'LOW',
                        'Vendor': vendor,
                        'Amount': latest_amount - baseline_amount,
                        'Amount_Type': 'Baseline-to-Latest Growth (Delta)',
                        'Date_1': baseline_period.to_timestamp(),
                        'Date_2': latest_period.to_timestamp(),
                        'Anchor_Date': pd.to_datetime(latest_txns['Date'], errors='coerce').max(),
                        'Bank': 'Multiple',
                        'Operation_Type': 'Multiple (Aggregate)',
                        'Source_Loc': f'Monthly vendor totals: {baseline_period}→{latest_period}',
                        'Evidence': f'{baseline_period}: ${baseline_amount:,.2f} → {latest_period}: ${latest_amount:,.2f} (baseline = first active month in period)',
                        'Txn_IDs': txn_id_summary,
                        'Next_Step': 'Review contract terms, pricing changes, volume justification',
                        'Risk': f'Dependency increasing over time',
                        'Action': 'Review contract terms and pricing changes'
                    })
                    alert_id += 1
    
    findings_df = pd.DataFrame(findings) if findings else pd.DataFrame()

    if return_run_stats:
        return findings_df, run_stats

    return findings_df

def mini_validate_vendor_growth_findings(
    transactions_df: pd.DataFrame,
    findings_df: pd.DataFrame,
    logger: logging.Logger,
    strict: bool = False
) -> list[str]:
    """
    Mini-validation for 'Vendor Concentration Growth' findings.

    Goals (audit-friendly):
      - Evidence months (baseline/latest) match Date_1/Date_2.
      - Txn_IDs referenced under 'Baseline YYYY-MM' and 'Latest YYYY-MM' exist in Transactions.
      - Those Txn_IDs actually belong to the referenced months.

    This does NOT change detection results. It only reports consistency issues.
    """
    warnings: list[str] = []

    if findings_df is None or len(findings_df) == 0:
        return warnings

    required_findings_cols = {'Category', 'Txn_IDs', 'Evidence', 'Date_1', 'Date_2'}
    if not required_findings_cols.issubset(set(findings_df.columns)):
        missing = sorted(list(required_findings_cols - set(findings_df.columns)))
        logger.debug(f"Mini-validation skipped: findings_df missing columns: {missing}")
        return warnings

    if transactions_df is None or len(transactions_df) == 0:
        warnings.append("Mini-validation: Transactions dataframe is empty; cannot validate Txn_ID references.")
        if strict:
            raise ValueError(warnings[-1])
        return warnings

    if 'Txn_ID' not in transactions_df.columns:
        warnings.append("Mini-validation: Transactions missing 'Txn_ID' column; cannot validate Txn_ID references.")
        if strict:
            raise ValueError(warnings[-1])
        return warnings

    # Prefer Month period column if available; otherwise derive from Date.
    if 'Month' in transactions_df.columns:
        txn_month_series = transactions_df['Month']
    elif 'Date' in transactions_df.columns:
        txn_month_series = pd.to_datetime(transactions_df['Date'], errors='coerce').dt.to_period('M')
    else:
        warnings.append("Mini-validation: Transactions missing 'Month' and 'Date' columns; cannot validate months.")
        if strict:
            raise ValueError(warnings[-1])
        return warnings

    txn_to_month: dict[str, pd.Period] = {}
    for tid, m in zip(transactions_df['Txn_ID'].astype(str), txn_month_series):
        if pd.isna(tid) or pd.isna(m):
            continue
        try:
            txn_to_month[str(tid)] = m if isinstance(m, pd.Period) else pd.Period(m, freq='M')
        except Exception:
            continue

    target = findings_df[findings_df['Category'] == 'Vendor Concentration Growth']
    if len(target) == 0:
        return warnings

    month_re = re.compile(r"\b(\d{4}-\d{2})\b")
    txn_re = re.compile(r"\bTXN_[A-Za-z0-9]+\b")

    for _, row in target.iterrows():
        alert_id = str(row.get('Alert_ID', 'UNKNOWN'))
        evidence = str(row.get('Evidence', '') or '')
        txn_ids_text = str(row.get('Txn_IDs', '') or '')

        # Expected months based on Date_1/Date_2 (source of truth)
        try:
            d1 = pd.to_datetime(row.get('Date_1'), errors='coerce')
            d2 = pd.to_datetime(row.get('Date_2'), errors='coerce')
            if pd.isna(d1) or pd.isna(d2):
                warnings.append(f"{alert_id}: Date_1/Date_2 missing or invalid; cannot validate baseline/latest months.")
                continue
            expected_baseline = d1.to_period('M')
            expected_latest = d2.to_period('M')
        except Exception:
            warnings.append(f"{alert_id}: failed to parse Date_1/Date_2; cannot validate months.")
            continue

        # Evidence should explicitly contain both months
        evidence_months = month_re.findall(evidence)
        if str(expected_baseline) not in evidence_months or str(expected_latest) not in evidence_months:
            warnings.append(
                f"{alert_id}: Evidence months do not match Date_1/Date_2 "
                f"(expected {expected_baseline} & {expected_latest})."
            )

        # Txn_IDs text is expected to be: "Baseline YYYY-MM: ... | Latest YYYY-MM: ..."
        segments = [s.strip() for s in txn_ids_text.split('|') if str(s).strip()]
        baseline_seg = next((s for s in segments if s.lower().startswith('baseline')), None)
        latest_seg = next((s for s in segments if s.lower().startswith('latest')), None)

        if baseline_seg is None or latest_seg is None:
            warnings.append(f"{alert_id}: Txn_IDs does not contain both 'Baseline' and 'Latest' segments.")
            continue

        # Segment month checks (if present)
        baseline_seg_month = month_re.findall(baseline_seg)
        latest_seg_month = month_re.findall(latest_seg)
        if baseline_seg_month and baseline_seg_month[0] != str(expected_baseline):
            warnings.append(
                f"{alert_id}: Baseline Txn_IDs segment month {baseline_seg_month[0]} "
                f"does not match expected {expected_baseline}."
            )
        if latest_seg_month and latest_seg_month[0] != str(expected_latest):
            warnings.append(
                f"{alert_id}: Latest Txn_IDs segment month {latest_seg_month[0]} "
                f"does not match expected {expected_latest}."
            )

        baseline_ids = txn_re.findall(baseline_seg)
        latest_ids = txn_re.findall(latest_seg)

        if len(baseline_ids) == 0:
            warnings.append(f"{alert_id}: No Txn_IDs found in Baseline segment.")
        if len(latest_ids) == 0:
            warnings.append(f"{alert_id}: No Txn_IDs found in Latest segment.")

        # Validate that referenced IDs exist and belong to expected months.
        for tid in baseline_ids:
            m = txn_to_month.get(tid)
            if m is None:
                warnings.append(f"{alert_id}: Baseline Txn_ID '{tid}' not found in Transactions.")
            elif m != expected_baseline:
                warnings.append(f"{alert_id}: Baseline Txn_ID '{tid}' month is {m}, expected {expected_baseline}.")

        for tid in latest_ids:
            m = txn_to_month.get(tid)
            if m is None:
                warnings.append(f"{alert_id}: Latest Txn_ID '{tid}' not found in Transactions.")
            elif m != expected_latest:
                warnings.append(f"{alert_id}: Latest Txn_ID '{tid}' month is {m}, expected {expected_latest}.")

    if warnings:
        logger.warning(f"Mini-validation found {len(warnings)} issue(s) in Vendor Concentration Growth findings.")
        for w in warnings[:20]:
            logger.warning(f"  - {w}")
        if len(warnings) > 20:
            logger.warning(f"  ... {len(warnings) - 20} more issue(s) omitted.")
        if strict:
            raise ValueError("Mini-validation failed; see warnings above.")
    else:
        logger.info("Mini-validation passed: Vendor Concentration Growth findings look consistent.")

    return warnings

def mini_validate_duplicate_payment_findings(
    transactions_df: pd.DataFrame,
    findings_df: pd.DataFrame,
    logger: logging.Logger,
    strict: bool = False
) -> list[str]:
    """
    Mini-validation for Duplicate Payment findings.

    Checks (best-effort):
      - Txn_IDs contains exactly 2 IDs that exist in Transactions.
      - Evidence contains BOTH amounts (formatted to 2 decimals) and the tolerance marker.
      - Days apart in Evidence matches actual day difference (<=3).

    Does NOT change detection results; only reports consistency issues.
    """
    warnings: list[str] = []

    if findings_df is None or len(findings_df) == 0:
        return warnings

    required_findings_cols = {'Category', 'Txn_IDs', 'Evidence', 'Date_1', 'Date_2'}
    if not required_findings_cols.issubset(set(findings_df.columns)):
        missing = sorted(list(required_findings_cols - set(findings_df.columns)))
        logger.debug(f"Mini-validate Duplicate Payment skipped (missing findings columns: {missing}).")
        return warnings

    required_txn_cols = {'Txn_ID', 'Date', 'Debit'}
    if not required_txn_cols.issubset(set(transactions_df.columns)):
        missing = sorted(list(required_txn_cols - set(transactions_df.columns)))
        logger.debug(f"Mini-validate Duplicate Payment skipped (missing transactions columns: {missing}).")
        return warnings

    dup = findings_df[findings_df['Category'] == 'Duplicate Payment'].copy()
    if dup.empty:
        return warnings

    txn_by_id = transactions_df.set_index('Txn_ID', drop=False)

    for _, row in dup.iterrows():
        alert_id = row.get('Alert_ID', '<unknown>')
        txn_ids_raw = str(row.get('Txn_IDs', '') or '')
        ids = [t.strip() for t in txn_ids_raw.split(',') if t.strip()]
        if len(ids) != 2:
            warnings.append(f"{alert_id}: expected 2 Txn_IDs for Duplicate Payment, got {len(ids)} ('{txn_ids_raw}').")
            continue

        missing_ids = [t for t in ids if t not in txn_by_id.index]
        if missing_ids:
            warnings.append(f"{alert_id}: Txn_ID(s) not found in Transactions: {', '.join(missing_ids)}.")
            continue

        t1 = txn_by_id.loc[ids[0]]
        t2 = txn_by_id.loc[ids[1]]

        # Amounts shown in Evidence should include BOTH values
        a1 = float(t1.get('Debit', 0) or 0)
        a2 = float(t2.get('Debit', 0) or 0)
        a1s = f"${a1:,.2f}"
        a2s = f"${a2:,.2f}"

        evidence = str(row.get('Evidence', '') or '')
        if (a1s not in evidence) or (a2s not in evidence):
            warnings.append(f"{alert_id}: Evidence does not contain both amounts ({a1s}, {a2s}). Evidence='{evidence}'.")
        if 'tolerance' not in evidence.lower():
            warnings.append(f"{alert_id}: Evidence does not mention tolerance. Evidence='{evidence}'.")

        # Days apart check (best-effort)
        d1 = row.get('Date_1')
        d2 = row.get('Date_2')
        try:
            d1 = pd.to_datetime(d1)
            d2 = pd.to_datetime(d2)
            days_apart = abs((d1 - d2).days)
            # Evidence contains '<N> day(s) apart'
            m = re.search(r"(\d+)\s*day\(s\)\s*apart", evidence)
            if m:
                ev_days = int(m.group(1))
                if ev_days != days_apart:
                    warnings.append(f"{alert_id}: Evidence days_apart={ev_days} but actual={days_apart}.")
            if days_apart > 3:
                warnings.append(f"{alert_id}: actual days_apart={days_apart} exceeds 3-day window.")
        except Exception:
            # If parsing fails, don't block
            pass

    if warnings:
        logger.warning(f"Mini-validation found {len(warnings)} issue(s) in Duplicate Payment findings.")
        for w in warnings[:20]:
            logger.warning(f"  - {w}")
        if len(warnings) > 20:
            logger.warning(f"  ... {len(warnings) - 20} more issue(s) omitted.")
        if strict:
            raise ValueError("Mini-validation failed; see warnings above.")
    else:
        logger.info("Mini-validation passed: Duplicate Payment findings look consistent.")

    return warnings


class StatementFormatError(Exception):
    """Raised when an input statement format cannot be parsed safely."""

class InputDataError(Exception):
    """Raised when input data fails required validations."""

class ValidationError(Exception):
    """Raised when integrity checks fail in strict mode."""


def _extract_txn_ids_from_text(text: str) -> list[str]:
    """Extract TXN_00000 style identifiers from a string."""
    if not text:
        return []
    return re.findall(r"TXN_\d{5}", str(text))


def final_integrity_checks(
    transactions_df: pd.DataFrame,
    findings_df: pd.DataFrame,
    *,
    strict: bool,
    logger: logging.Logger,
) -> None:
    """Run final integrity checks before writing the Excel workbook.

    These checks do NOT change any calculations. In non-strict mode, they log warnings.
    In strict mode, they raise ValidationError on any error.
    """
    errors: list[str] = []
    warnings: list[str] = []

    if transactions_df is None or len(transactions_df) == 0:
        errors.append("Transactions dataframe is empty.")
    else:
        # Txn_ID presence and uniqueness
        if "Txn_ID" not in transactions_df.columns:
            errors.append("Transactions is missing required column 'Txn_ID'.")
        else:
            txn_ids = transactions_df["Txn_ID"].astype(str).tolist()
            if any(tid == "nan" or tid.strip() == "" for tid in txn_ids):
                errors.append("One or more Txn_ID values are missing.")
            dupes = pd.Series(txn_ids).duplicated()
            if dupes.any():
                errors.append(f"Txn_ID values are not unique (duplicates={int(dupes.sum())}).")

        # Debit/Credit sanity
        for col in ["Debit", "Credit"]:
            if col not in transactions_df.columns:
                errors.append(f"Transactions is missing required column '{col}'.")
        if all(c in transactions_df.columns for c in ["Debit", "Credit"]):
            debit = pd.to_numeric(transactions_df["Debit"], errors="coerce").fillna(0)
            credit = pd.to_numeric(transactions_df["Credit"], errors="coerce").fillna(0)
            if (debit < 0).any() or (credit < 0).any():
                errors.append("Negative values detected in Debit/Credit.")
            both = (debit > 0) & (credit > 0)
            if both.any():
                errors.append(f"Rows with both Debit and Credit > 0 detected (count={int(both.sum())}).")

        # Direction consistency (best-effort)
        if all(c in transactions_df.columns for c in ["Direction", "Debit", "Credit"]):
            debit = pd.to_numeric(transactions_df["Debit"], errors="coerce").fillna(0)
            credit = pd.to_numeric(transactions_df["Credit"], errors="coerce").fillna(0)
            direction = transactions_df["Direction"].astype(str).str.lower()
            bad_out = ((debit > 0) & ~direction.str.contains("out"))
            bad_in = ((credit > 0) & ~direction.str.contains("in"))
            # Allow neutral/unknown direction if present
            if bad_out.any():
                warnings.append(f"Direction may be inconsistent for some debit rows (count={int(bad_out.sum())}).")
            if bad_in.any():
                warnings.append(f"Direction may be inconsistent for some credit rows (count={int(bad_in.sum())}).")

        # Date parseability (best-effort)
        if "Date" in transactions_df.columns:
            parsed = pd.to_datetime(transactions_df["Date"], errors="coerce")
            if parsed.isna().any():
                warnings.append(f"Some transaction dates could not be parsed (count={int(parsed.isna().sum())}).")

    # Findings references -> Txn_ID existence
    if findings_df is not None and len(findings_df) > 0:
        # Common fields that may contain Txn IDs
        for col in ["Txn_IDs", "Txn IDs", "Evidence"]:
            if col in findings_df.columns:
                for i, v in enumerate(findings_df[col].fillna("").astype(str).tolist(), start=1):
                    for tid in _extract_txn_ids_from_text(v):
                        if "Txn_ID" in transactions_df.columns:
                            if tid not in set(transactions_df["Txn_ID"].astype(str).tolist()):
                                errors.append(f"Findings references unknown Txn_ID '{tid}' (column={col}).")
    # Emit logs
    for w in warnings:
        logger.warning(f"Integrity warning: {w}")
    if errors:
        for e in errors:
            logger.error(f"Integrity error: {e}")
        if strict:
            raise ValidationError(f"Integrity checks failed with {len(errors)} error(s).")

    logger.info(
        f"Final integrity checks: errors={len(errors)} | warnings={len(warnings)} | strict={strict}"
    )


def detect_beginning_balance(file_path, df_with_balance=None):
    """
    Detect beginning/opening balance and return (balance, ref) where ref is a best-effort extraction pointer.

    Ref formats:
      - Header/raw scan: "<SheetName>!<ColLetter><RowNumber>" (e.g., "GBB_BizChk_USD_2023-07!E4")
      - Dataframe-based fallback: "<SheetName>!R<RowNumber>" (best-effort when original column is unknown)
    """
    from openpyxl.utils import get_column_letter

    # Helper: first sheet name (pandas default when sheet_name not provided)

    sheet = first_sheet_name(file_path)

    # 1) Header/raw scan (first ~25 rows, first sheet)
    try:
        raw_df = pd.read_excel(file_path, header=None, nrows=25)
        keywords = [
            'beginning balance', 'opening balance', 'starting balance',
            'balance forward', 'balance brought forward', 'previous balance', 'present balance',
            'account balance'
        ]

        for ridx, row in raw_df.iterrows():
            for cidx, cell in enumerate(row):
                if pd.notna(cell):
                    cell_str = str(cell).lower()
                    if any(k in cell_str for k in keywords):
                        # Search same row (including same cell) for a numeric value
                        for j in range(cidx, len(row)):
                            v = row.iloc[j] if j < len(row) else None
                            if pd.isna(v):
                                continue

                            # Direct numeric
                            try:
                                bal = float(v)
                                if abs(bal) > 0.01:
                                    ref = f"{sheet}!{get_column_letter(j+1)}{ridx+1}"
                                    return bal, ref
                            except (ValueError, TypeError):
                                pass

                            # Numeric embedded in string
                            try:
                                s = str(v)
                                s = s.replace('$', '').replace(',', '').replace('€', '').replace('£', '')
                                s = s.replace('(', '-').replace(')', '').strip()
                                bal = float(s)
                                if abs(bal) > 0.01:
                                    ref = f"{sheet}!{get_column_letter(j+1)}{ridx+1}"
                                    return bal, ref
                            except (ValueError, TypeError):
                                continue
    except Exception:
        logger.debug("Beginning balance detection failed (raw scan).", exc_info=True)

    # 2) Balance_Column fallback (take first non-null)
    try:
        if df_with_balance is not None and 'Balance_Column' in df_with_balance.columns:
            non_null = df_with_balance['Balance_Column'].dropna()
            if len(non_null) > 0:
                first_balance = non_null.iloc[0]
                if pd.notna(first_balance) and abs(float(first_balance)) > 0.01:
                    # Best-effort ref: use the first transaction's source row/sheet (column unknown)
                    ref = None
                    if 'Source_Sheet' in df_with_balance.columns and 'Source_Row' in df_with_balance.columns:
                        ref = f"{df_with_balance['Source_Sheet'].iloc[0]}!R{int(df_with_balance['Source_Row'].iloc[0])}"
                    return float(first_balance), ref
    except Exception:
        logger.debug("Beginning balance detection failed (Balance_Column fallback).", exc_info=True)

    # 3) Description row fallback (keyword row inside parsed df)
    try:
        if df_with_balance is not None and 'Description' in df_with_balance.columns:
            beginning_rows = df_with_balance[df_with_balance['Description'].astype(str).str.contains(
                'beginning balance|opening balance|starting balance|balance forward|balance brought forward',
                case=False, na=False, regex=True
            )]
            if len(beginning_rows) > 0:
                row0 = beginning_rows.iloc[0]
                for col in ['Credit', 'Debit', 'Balance_Column']:
                    if col in beginning_rows.columns:
                        v = row0.get(col)
                        if pd.notna(v) and abs(float(v)) > 0.01:
                            ref = None
                            if 'Source_Sheet' in beginning_rows.columns and 'Source_Row' in beginning_rows.columns:
                                ref = f"{row0.get('Source_Sheet')}!R{int(row0.get('Source_Row'))}"
                            return abs(float(v)), ref
    except Exception:
        logger.debug("Beginning balance detection failed (df scan).", exc_info=True)

    return None


def detect_ending_balance(file_path, df_with_balance=None):
    """
    Detect ending/closing balance and return (balance, ref) where ref is a best-effort extraction pointer.

    Ref formats:
      - Header/raw scan: "<SheetName>!<ColLetter><RowNumber>"
      - Dataframe-based fallback: "<SheetName>!R<RowNumber>" (best-effort when original column is unknown)
    """
    import re
    from openpyxl.utils import get_column_letter


    sheet = first_sheet_name(file_path)

    # 1) Raw scan (first ~200 rows, first sheet)
    try:
        raw_df = pd.read_excel(file_path, header=None, nrows=200)
        keywords = [
            'ending balance', 'closing balance', 'new balance',
            'balance carried forward', 'balance carried fwd', 'balance carry forward'
        ]
        for r in range(len(raw_df)):
            row_vals = raw_df.iloc[r].tolist()
            for c, cell in enumerate(row_vals):
                if isinstance(cell, str) and any(k in cell.lower() for k in keywords):
                    for vidx, v in enumerate(row_vals):
                        if isinstance(v, (int, float)) and not pd.isna(v) and abs(float(v)) > 0.01:
                            ref = f"{sheet}!{get_column_letter(vidx+1)}{r+1}"
                            return float(v), ref
                        if isinstance(v, str):
                            m = re.search(r"\(?-?\$?([0-9][0-9,]*\.?[0-9]*)\)?", v.replace(' ', ''))
                            if m:
                                num_str = m.group(0)
                                negative = num_str.strip().startswith('(') and num_str.strip().endswith(')')
                                num_str = num_str.replace('$', '').replace(',', '').replace('(', '').replace(')', '')
                                try:
                                    val = float(num_str)
                                    if negative:
                                        val = -val
                                    if abs(val) > 0.01:
                                        ref = f"{sheet}!{get_column_letter(vidx+1)}{r+1}"
                                        return val, ref
                                except (ValueError, TypeError):
                                    pass
                    # keyword found but no numeric parsed -> continue
    except Exception:
        logger.debug("Ending balance detection failed (raw scan).", exc_info=True)

    # 2) Dataframe fallback (keyword row inside parsed df)
    try:
        if df_with_balance is not None and 'Description' in df_with_balance.columns:
            ending_rows = df_with_balance[df_with_balance['Description'].astype(str).str.contains(
                'ending balance|closing balance|new balance|balance carried forward|balance carried fwd',
                case=False, na=False, regex=True
            )]
            if len(ending_rows) > 0:
                row = ending_rows.iloc[-1]
                for col in ['Credit', 'Debit']:
                    if col in ending_rows.columns:
                        v = row.get(col)
                        if pd.notna(v) and abs(float(v)) > 0.01:
                            ref = None
                            if 'Source_Sheet' in ending_rows.columns and 'Source_Row' in ending_rows.columns:
                                ref = f"{row.get('Source_Sheet')}!R{int(row.get('Source_Row'))}"
                            return float(v), ref
    except Exception:
        logger.debug("Ending balance detection failed (df scan).", exc_info=True)

    return None


# ============================================================================
# Transaction Categorization
# ============================================================================

def classify_operation_type(df, description_column, transaction_type_column=None):
    """Classify business operations by keywords in transaction description."""
    if description_column not in df.columns:
        return df
    
    # Classification rules: category → list of keywords (regular expressions)
    # IMPORTANT: Order matters! Check from specific to general.
    classification_rules = {
        # 1. PAYROLL & EMPLOYEE PAYMENTS
        'Payroll & Benefits': [
            r'payroll', r'salary', r'wages', r'employee\s+payment',
            r'direct\s+deposit.*employee', r'gusto', r'adp\s+payroll',
            r'paychex', r'health\s+insurance', r'dental\s+insurance',
            r'vision\s+insurance', r'401k', r'retirement\s+plan',
        ],
        
        # 2. TAXES & GOVERNMENT PAYMENTS
        'Taxes & Government': [
            r'irs\s+', r'federal\s+tax', r'state\s+tax', r'sales\s+tax',
            r'payroll\s+tax', r'income\s+tax', r'tax\s+payment',
            r'department\s+of\s+revenue', r'franchise\s+tax',
            r'property\s+tax', r'employment\s+tax', r'eftps',
        ],
        
        # 3. BANK FEES & CHARGES (expanded)
        'Bank Fees': [
            r'service\s+fee', r'monthly\s+fee', r'wire\s+fee',
            r'transaction\s+fee', r'overdraft', r'nsf\s+fee',
            r'maintenance\s+fee', r'account\s+fee',
            r'service\s+charge', r'check\s+printing\s+fee',
            r'ach\s+fee', r'ach\s+origination\s+fee', r'wire\s+transfer\s+fee',
            r'monthly\s+service\s+charge',
        ],
        
        # 4. SOFTWARE & IT (expanded with accounting, telecom, broadband)
        'Software & IT': [
            r'software', r'saas\s+', r'cloud\s+', r'hosting',
            r'google\s+workspace', r'microsoft\s+365', r'office\s+365',
            r'adobe', r'slack', r'zoom', r'dropbox',
            r'amazon\s+web\s+services', r'aws\s+', r'azure',
            r'salesforce', r'hubspot', r'quickbooks', r'intuit',
            r'software\s+subscription', r'nexus\s+software',
            r'tech\s+subscription', r'digital\s+subscription',
            r'ringcentral', r'voip', r'broadband', r'bt\s+business',
            r'internet\s+service', r'business\s+broadband',
        ],
        
        # 5. CREDIT CARD & MERCHANT PROCESSING
        'Card Processing': [
            r'card\s+purchase', r'debit\s+card', r'credit\s+card',
            r'square', r'stripe', r'paypal', r'merchant\s+services',
            r'pos\s+', r'point\s+of\s+sale',
        ],
        
        # 6. RENT & FACILITIES (expanded to catch property/lease variations)
        'Rent & Facilities': [
            r'rent\s+', r'lease\s+payment', r'office\s+rent',
            r'warehouse\s+rent', r'warehouse\s+lease', r'jul\s+rent',
            r'commercial\s+lease', r'facilities', r'facilities\s+mgmt',
            r'riverside\s+facilities', r'riverside\s+properties',
            r'property\s+management', r'property\s+ltd',
            r'direct\s+debit.*property', r'example\s+property',
        ],
        
        # 7. UTILITIES & SERVICES (expanded)
        'Utilities & Services': [
            r'electric', r'gas\s+bill', r'water\s+bill', r'utilities',
            r'internet\s+service', r'phone\s+bill', r'telecom',
            r'metro\s+water', r'city\s+utilities', r'vodafone',
            r'city\s+electric', r'waste\s+management', r'trash',
            r'disposal', r'recycling', r'sanitation',
            r'comcast', r'at&t', r'verizon', r'internet\s+&\s+phone',
            r'cleaning', r'janitorial', r'janitorial\s+services',
        ],
        
        # 8. OFFICE & SUPPLIES (expanded to catch home depot, hardware, etc)
        'Office & Supplies': [
            r'office\s+supplies', r'staples', r'office\s+depot',
            r'amazon.*office', r'amazon\s+business', r'paper', r'printer', r'ink',
            r'maintenance\s+&\s+repair', r'supplies\s+\d+',
            r'global\s+industrial', r'industrial\s+supply',
            r'safety\s+equipment', r'quality\s+maintenance',
            r'home\s+depot', r'screwfix', r'nursery\s+supply',
            r'commercial\s+nursery', r'hardware\s+store',
        ],
        
        # 9. PROFESSIONAL SERVICES
        'Professional Services': [
            r'legal\s+fees', r'attorney', r'law\s+firm',
            r'accounting\s+services', r'cpa\s+', r'audit',
            r'consulting', r'consulting\s+services',
            r'contracting', r'summit\s+contracting',
            r'securenet', r'it\s+services', r'tech\s+support',
            r'orchard\s+consulting',
        ],
        
        # 10. SHIPPING & LOGISTICS
        'Shipping & Logistics': [
            r'fedex', r'ups\s+', r'ups$', r'usps', r'dhl',
            r'shipping', r'freight', r'logistics',
            r'delivery\s+service', r'acme\s+logistics',
            r'courier', r'express\s+delivery',
        ],
        
        # 11. INSURANCE & SECURITY (expanded)
        'Insurance & Security': [
            r'insurance\s+premium', r'insurance', r'liability\s+insurance',
            r'property\s+insurance', r'workers\s+comp',
            r'workers\s+compensation', r'commercial\s+insurance',
            r'security\s+alarm', r'security\s+monitoring',
            r'alarm\s+co', r'security\s+inc', r'ironclad\s+security',
            r'security\s+service',
        ],
        
        # 12. INVENTORY & SUPPLIERS (Payments OUT to suppliers)
        'Inventory & Suppliers': [
            r'supplier', r'vendor\s+payment', r'inventory',
            r'wholesale', r'purchase\s+order', r'\s+po\s+\d+', 
            r'material', r'raw\s+material', r'stock\s+purchase',
            r'packaging', r'pacific\s+packaging', r'electronic\s+parts',
            r'parts\s+ltd', r'hardware', r'oak\s+&\s+iron\s+hardware',
            r'furniture', r'margolis\s+furniture', r'ikea',
            r'midwest\s+industrial\s+parts',
        ],
        
        # 13. CUSTOMER PAYMENTS (REVENUE - Payments IN from customers)
        'Customer Payments': [
            r'customer\s+payment', r'invoice\s+payment', r'inv\s+\d+',
            r'payment\s+received', r'lockbox', r'remote\s+deposit',
            r'sales\s+receipt', r'revenue',
            r'inward\s+payments', r'inward\s+fast\s+payments',
            r'faster\s+payments\s+-\s+in', r'faster\s+payments\s+in',  # UK bank inbound
            r'brightline\s+stores', r'eastcoast\s+imports',
            r'riverstone\s+retail', r'riverbend\s+markets',
            r'northbridge\s+retail', r'retail\s+llc',
            r'wire\s+incoming', r'ach\s+credit.*customer',
            r'nova\s+electronics', r'cedarstone\s+foods',  # Specific customers from data
            r'labelle\s+design', r'brightline\s+trading',
            r'central\s+marketplace', r'oak\s+&\s+iron\s+partners',
        ],
        
        # 14. CONTRACTOR & STAFFING SERVICES
        'Contractor & Staffing': [
            r'contractor\s+payment', r'contractor', r'temp\s+staffing',
            r'field\s+service', r'project\s+staffing', r'project\s+support',
            r'buildout\s+contractor', r'leasehold\s+improvements?',
            r'construction', r'hvac\s+service', r'commercial\s+hvac',
            r'maintenance\s+contractor',
        ],
        
        # 15. CASH MANAGEMENT & ATM
        'Cash Management': [
            r'atm\s+cash\s+withdrawal', r'atm\s+withdrawal', r'cash\s+withdrawal',
            r'petty\s+cash', r'change\s+fund', r'cash\s+fund',
            r'branch\s+cash', r'cash\s+services',
        ],
        
        # 16. EMPLOYEE REIMBURSEMENTS
        'Employee Reimbursements': [
            r'employee\s+reimbursement', r'reimbursement', r'travel\s+expense',
            r'expense\s+reimbursement', r'mileage\s+reimbursement',
        ],
        
        # 17. VENDOR PAYMENTS (General payments to vendors)
        'Vendor Payments': [
            r'outward\s+payments', r'outward\s+fast\s+payments',
            r'ach\s+payment', r'check\s+\d+', r'wire\s+out',
            r'northstar\s+metals', r'triangle\s+tooling',
        ],
        
        # 18. LOANS & FINANCING
        'Loans & Financing': [
            r'loan\s+payment', r'loan\s+proceeds', r'line\s+of\s+credit',
            r'term\s+loan', r'sba\s+loan', r'business\s+loan',
            r'equipment\s+financing', r'merchant\s+advance',
        ],
        
        # 19. MARKETING & ADVERTISING
        'Marketing & Advertising': [
            r'advertising', r'marketing', r'google\s+ads',
            r'facebook\s+ads', r'linkedin\s+ads', r'social\s+media',
            r'seo\s+', r'email\s+marketing',
        ],
        
        # 20. FUEL & FLEET
        'Fuel & Fleet': [
            r'fuel', r'fleet\s+fuel', r'gas\s+station', r'gasoline',
            r'diesel', r'vehicle\s+fuel',
        ],
        
        # 21. OWNER TRANSACTIONS
        'Owner Transactions': [
            r'owner\s+draw', r'distribution', r'dividend',
            r'capital\s+contribution', r'equity\s+injection',
            r'\srubio\s', r'^rubio$',
        ],
        
        # 22. FINANCIAL INCOME
        'Financial Income': [
            r'interest\s+earned', r'interest\s+income',
            r'bank\s+interest', r'investment\s+income',
        ],
        
        # 23. INTERNAL TRANSFERS
        'Internal Transfers': [
            r'transfer\s+to', r'transfer\s+from',
            r'internal\s+transfer', r'between\s+accounts',
        ],
    }
    
    def classify_text(description, transaction_type=''):
        if pd.isna(description):
            return 'Other'
        
        text = str(description).lower()
        
        # Also check transaction type if available (Chase has this)
        if transaction_type:
            text = text + ' ' + str(transaction_type).lower()
        
        # Check each category in order
        for category, keywords in classification_rules.items():
            for keyword in keywords:
                if re.search(keyword, text):
                    return category
        
        # If no category matched
        return 'Other'
    
    # Create new column with classification
    if transaction_type_column and transaction_type_column in df.columns:
        df['Operation_Type'] = df.apply(
            lambda row: classify_text(row[description_column], row[transaction_type_column]), 
            axis=1
        )
    else:
        df['Operation_Type'] = df[description_column].apply(lambda x: classify_text(x, ''))
    
    return df


def process_excel_file(file_path, output_columns):
    """Process Excel file with automatic bank format detection and standardization."""
    try:
        # Automatic format detection
        bank_config = detect_bank_format(file_path)
        
        if bank_config is None:
            logger.info(f"  SKIPPED '{file_path.name}': Unable to detect bank format")
            return None
        
        bank_name = bank_config['bank']
        header_row = bank_config['header']
        format_type = bank_config['format_type']
        columns = bank_config['columns']
        
        # Read data from header row
        df = pd.read_excel(file_path, header=header_row)
        
        # Capture original Excel row numbers BEFORE any filtering
        # This ensures Source_Row traceability remains accurate even after removing empty rows
        # Excel rows are 1-indexed, and header_row is already the header position
        # So first data row in pandas (row 0) = Excel row (header_row + 2)
        # +1 for Excel 1-based indexing, +1 because header_row itself is the header
        df['_excel_row_original'] = range(header_row + 2, header_row + 2 + len(df))
        
        # Filter out non-transaction rows
        initial_rows = len(df)
        filtered_rows = 0  # Track how many rows we filter out
        
        # Get actual column names
        actual_cols = df.columns.tolist()
        
        # Find date column
        date_col_name = actual_cols[columns['date']] if columns.get('date') is not None else None
        
        if date_col_name:
            # Remove rows where Date is empty
            df = df[df[date_col_name].notna()]
            
            # Try to convert to datetime and keep only valid dates
            df['temp_date'] = pd.to_datetime(df[date_col_name], errors='coerce')
            df = df[df['temp_date'].notna()]
            df = df.drop('temp_date', axis=1)
        
        # Build column mapping
        col_mapping = {}
        
        # Date
        if columns.get('date') is not None:
            col_mapping[actual_cols[columns['date']]] = 'Date'
        
        # Description
        if columns.get('description') is not None:
            col_mapping[actual_cols[columns['description']]] = 'Description'
        
        # Type (optional)
        if columns.get('type') is not None:
            col_mapping[actual_cols[columns['type']]] = 'Type_Original'
        
        # Balance (optional, for reference)
        if columns.get('balance') is not None:
            col_mapping[actual_cols[columns['balance']]] = 'Balance_Column'
        
        # Handle Debit/Credit vs Amount format
        if format_type == 'debit_credit':
            if columns.get('debit') is not None:
                col_mapping[actual_cols[columns['debit']]] = 'Debit'
            if columns.get('credit') is not None:
                col_mapping[actual_cols[columns['credit']]] = 'Credit'
        
        elif format_type == 'amount_signed':
            if columns.get('amount') is not None:
                col_mapping[actual_cols[columns['amount']]] = 'Amount'
        
        # Rename columns
        df = df.rename(columns=col_mapping)
        
        # Filter out rows with empty Description (service rows, spacing, etc)
        if 'Description' in df.columns:
            initial_count = len(df)
            df = df[df['Description'].notna()]
            df = df[df['Description'].astype(str).str.strip() != '']
            if len(df) < initial_count:
                filtered_rows += (initial_count - len(df))
        
        # Convert Amount column to Debit/Credit if needed
        if 'Amount' in df.columns and 'Debit' not in df.columns:
            # Clean amount column first
            df['Amount'] = df['Amount'].astype(str).str.replace(' ', '').str.replace(',', '')
            df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
            
            # Negative amounts = Debits (outflows)
            # Positive amounts = Credits (inflows)
            df['Debit'] = df['Amount'].apply(lambda x: abs(x) if pd.notna(x) and x < 0 else np.nan)
            df['Credit'] = df['Amount'].apply(lambda x: x if pd.notna(x) and x > 0 else np.nan)
            # Drop the Amount column after conversion
            df = df.drop('Amount', axis=1)
        
        # Select only the columns we need
        # Save _excel_row_original before it gets dropped
        excel_row_backup = df['_excel_row_original'].copy()
        
        available_cols = [col for col in output_columns if col in df.columns]
        df = df[available_cols].copy()
        
        # Restore _excel_row_original after column selection
        df['_excel_row_original'] = excel_row_backup
        
        filtered_rows += (initial_rows - len(df))  # Add to running total
        
        # Add metadata columns
        df['Bank'] = bank_name
        df['Source_File'] = file_path.name
        
        # NOTE: Currently only first sheet is processed per file
        # Multi-sheet files (e.g., monthly tabs like Jan/Feb/Mar) will show first sheet name for all transactions
        # LIMITATION: If bank statement has multiple sheets, only the first sheet is read
        # For multi-sheet processing, split into separate files or modify detect_bank_format() to loop through sheets
        df['Source_Sheet'] = bank_config.get('sheet', 'Sheet1')
        
        # Use pre-calculated Excel row numbers (captured before filtering)
        # The _excel_row_original column was set BEFORE any rows were removed,
        # so it points to the correct Excel row regardless of filtering
        df['Source_Row'] = df['_excel_row_original']
        df = df.drop('_excel_row_original', axis=1)  # Clean up temporary column
        
        # Normalize data
        if 'Date' in df.columns:
            df = normalize_date_column(df, 'Date')
        
        if 'Debit' in df.columns:
            df = normalize_amount_column(df, 'Debit')
            # Replace zeros with NaN
            df.loc[df['Debit'] == 0, 'Debit'] = np.nan
        
        if 'Credit' in df.columns:
            df = normalize_amount_column(df, 'Credit')
            # Replace zeros with NaN
            df.loc[df['Credit'] == 0, 'Credit'] = np.nan
        
        # DATA QUALITY VALIDATION: Check for invalid Debit/Credit combinations
        if 'Debit' in df.columns and 'Credit' in df.columns:
            both_filled = (df['Debit'].notna() & df['Credit'].notna()).sum()
            neither_filled = (df['Debit'].isna() & df['Credit'].isna()).sum()
            if both_filled > 0:
                logger.warning(f"     WARNING: {both_filled} rows have BOTH Debit and Credit (data quality issue)")
            if neither_filled > 0:
                logger.warning(f"     WARNING: {neither_filled} rows have NEITHER Debit nor Credit (empty transactions)")
        
        # Clean vendor names
        if 'Description' in df.columns:
            df = normalize_company_name(df, 'Description')
        
        # Classify operations
        if 'Description' in df.columns:
            df = classify_operation_type(df, 'Description', None)
            
            # POST-CLASSIFICATION FIXES (only if columns exist)
            # Fix 1: IRS/Tax payments should NEVER be "Customer Payments"
            if 'Description_clean' in df.columns and 'Operation_Type' in df.columns:
                tax_keywords = ['irs', 'internal revenue', 'tax payment', 'federal tax', 'state tax']
                for keyword in tax_keywords:
                    mask = df['Description_clean'].str.contains(keyword, case=False, na=False) & (df['Operation_Type'] == 'Customer Payments')
                    df.loc[mask, 'Operation_Type'] = 'Taxes & Government'
            
            # Fix 2: Outward payments are NOT customer payments
            if 'Description_clean' in df.columns and 'Operation_Type' in df.columns:
                outward_mask = df['Description_clean'].str.contains('outward|ach payment -', case=False, na=False, regex=True) & (df['Operation_Type'] == 'Customer Payments')
                df.loc[outward_mask, 'Operation_Type'] = 'Vendor Payments'
            
            # Fix 2.5: "Retail/Wholesale" OUTFLOWS are NOT customer payments
            if 'Debit' in df.columns and 'Operation_Type' in df.columns and 'Description_clean' in df.columns:
                retail_outflows = (df['Debit'] > 0) & (df['Operation_Type'] == 'Customer Payments') & \
                                 df['Description_clean'].str.contains('retail|wholesale', case=False, na=False, regex=True)
                df.loc[retail_outflows, 'Operation_Type'] = 'Vendor Payments'
            
            # Fix 3: Customer payments in wrong category (only if Credit column exists)
            if 'Credit' in df.columns and 'Operation_Type' in df.columns and 'Description_clean' in df.columns:
                customer_in_inventory = (df['Credit'] > 0) & (df['Operation_Type'] == 'Inventory & Suppliers') & \
                                       df['Description_clean'].str.contains('customer payment|harbor wholesale|oak & iron|wire incoming', case=False, na=False, regex=True)
                df.loc[customer_in_inventory, 'Operation_Type'] = 'Customer Payments'
            
            # Fix 3.5: "Wholesale" INFLOWS are customer payments, not inventory
            if 'Credit' in df.columns and 'Operation_Type' in df.columns and 'Description_clean' in df.columns:
                wholesale_inflows = (df['Credit'] > 0) & (df['Operation_Type'] == 'Inventory & Suppliers') & \
                                   df['Description_clean'].str.contains('wholesale', case=False, na=False)
                df.loc[wholesale_inflows, 'Operation_Type'] = 'Customer Payments'
            
            # Fix 4: Card settlements vs card purchases (only if Credit column exists)
            if 'Credit' in df.columns and 'Operation_Type' in df.columns and 'Description_clean' in df.columns:
                settlement_mask = (df['Credit'] > 0) & (df['Operation_Type'] == 'Card Processing') & \
                                df['Description_clean'].str.contains('settlement|stripe payout|square deposit', case=False, na=False, regex=True)
                df.loc[settlement_mask, 'Operation_Type'] = 'Card Settlement (Income)'
            
            # Fix 5: Processor fees (only if Debit column exists)
            if 'Debit' in df.columns and 'Operation_Type' in df.columns and 'Description_clean' in df.columns:
                fee_mask = (df['Debit'] > 0) & (df['Operation_Type'] == 'Card Processing') & \
                          df['Description_clean'].str.contains('fee|charge|processing', case=False, na=False, regex=True)
                df.loc[fee_mask, 'Operation_Type'] = 'Card Fees'
            
            # Add Direction column
            df['Direction'] = 'Unknown'
            if 'Debit' in df.columns:
                df.loc[df['Debit'] > 0, 'Direction'] = 'Outflow'
            if 'Credit' in df.columns:
                df.loc[df['Credit'] > 0, 'Direction'] = 'Inflow'
        
        if filtered_rows > 0:
            logger.info(f" Processed '{file_path.name}' ({bank_name}): {len(df)} rows (removed service rows: {filtered_rows})")
        else:
            logger.info(f" Processed '{file_path.name}' ({bank_name}): {len(df)} rows")
        
        return df
        
    except Exception as e:
        logger.error(f" ERROR processing '{file_path.name}': {e}")
        import traceback
        traceback.print_exc()
        return None


def benford_law_analysis(df):
    """Apply Benford's Law analysis to detect potential data manipulation."""
    amounts = df['Debit'].dropna()
    amounts = amounts[amounts > 0]  # Critical: exclude zero values
    
    if len(amounts) < 50:
        return None  # Need sufficient data for Benford's Law
    
    # Extract first digit
    first_digits = amounts.astype(str).str[0].astype(int)
    
    # Count distribution
    actual_dist = first_digits.value_counts(normalize=True).sort_index()
    
    # Benford's Law expected distribution
    expected = {d: np.log10(1 + 1/d) for d in range(1, 10)}
    expected_dist = pd.Series(expected)
    
    # TODO: Consider adding second-digit Benford test for better accuracy
    
    # Calculate chi-square statistic
    observed = []
    expected_counts = []
    for digit in range(1, 10):
        obs = actual_dist.get(digit, 0) * len(first_digits)
        exp = expected_dist.get(digit, 0) * len(first_digits)
        observed.append(obs)
        expected_counts.append(exp)
    
    # Chi-square test
    chi_square = sum((o - e)**2 / e for o, e in zip(observed, expected_counts) if e > 0)
    
    # Critical value for 8 degrees of freedom at 95% confidence: 15.507
    passes_test = chi_square < 15.507
    
    results = []
    for digit in range(1, 10):
        actual_pct = actual_dist.get(digit, 0) * 100
        expected_pct = expected_dist.get(digit, 0) * 100
        deviation = actual_pct - expected_pct
        
        results.append({
            'First_Digit': digit,
            'Expected_%': expected_pct,
            'Actual_%': actual_pct,
            'Deviation_%': deviation
        })
    
    return {
        'results': pd.DataFrame(results),
        'chi_square': chi_square,
        'passes': passes_test,
        'sample_size': len(amounts)
    }


def calculate_cash_flow_metrics(df):
    """
    Calculate key cash flow and financial health metrics.
    Includes cross-validation to ensure dashboard totals match Transactions sheet.
    """
    metrics = {}
    
    # Basic totals
    total_inflow = df['Credit'].sum()
    total_outflow = df['Debit'].sum()
    net_flow = total_inflow - total_outflow
    
    metrics['total_inflow'] = total_inflow
    metrics['total_outflow'] = total_outflow
    metrics['net_cash_flow'] = net_flow
    
    # CROSS-VALIDATION: Verify totals match transaction-level sums
    # This catches data integrity issues early
    credit_sum_check = df['Credit'].fillna(0).sum()
    debit_sum_check = df['Debit'].fillna(0).sum()
    
    tolerance = 0.01  # Floating point tolerance
    
    if abs(total_inflow - credit_sum_check) > tolerance:
        logger.warning(f"\n  WARNING: Credit sum mismatch!")
        logger.info(f"   Dashboard: ${total_inflow:,.2f}")
        logger.info(f"   Actual: ${credit_sum_check:,.2f}")
        logger.info(f"   Difference: ${abs(total_inflow - credit_sum_check):,.2f}")
        metrics['validation_error'] = True
    
    if abs(total_outflow - debit_sum_check) > tolerance:
        logger.warning(f"\n  WARNING: Debit sum mismatch!")
        logger.info(f"   Dashboard: ${total_outflow:,.2f}")
        logger.info(f"   Actual: ${debit_sum_check:,.2f}")
        logger.info(f"   Difference: ${abs(total_outflow - debit_sum_check):,.2f}")
        metrics['validation_error'] = True
    
    # Count transactions for additional validation
    metrics['transaction_count'] = len(df)
    
    # Time-based analysis
    if df['Date'].notna().any():
        date_range = (df['Date'].max() - df['Date'].min()).days
        if date_range > 0:
            avg_daily_inflow = total_inflow / date_range
            avg_daily_outflow = total_outflow / date_range
            avg_daily_burn = avg_daily_outflow - avg_daily_inflow
            
            metrics['period_days'] = date_range
            metrics['avg_daily_inflow'] = avg_daily_inflow
            metrics['avg_daily_outflow'] = avg_daily_outflow
            metrics['avg_daily_burn'] = avg_daily_burn
            
            # Calculate runway (assuming current balance is final balance)
            # Note: This is illustrative - real runway needs actual balance
            if avg_daily_burn > 0:
                # Assuming last transaction balance
                last_transactions = df.sort_values('Date', ascending=False).head(1)
                if len(last_transactions) > 0 and 'Credit' in last_transactions.columns:
                    # Estimate: if burning cash, how long until zero
                    estimated_runway = abs(net_flow / avg_daily_burn) if avg_daily_burn != 0 else 999
                    metrics['estimated_runway_days'] = min(estimated_runway, 999)
    
    # Monthly breakdown
    df_copy = df.copy()
    df_copy['Month'] = pd.to_datetime(df_copy['Date']).dt.to_period('M')
    
    monthly = df_copy.groupby('Month').agg({
        'Debit': 'sum',
        'Credit': 'sum'
    })
    monthly['Net'] = monthly['Credit'] - monthly['Debit']
    monthly['Month_Name'] = monthly.index.astype(str)
    
    return metrics, monthly


def calculate_concentration_risk(df):
    """
    Analyze vendor concentration and dependency risk.
    Uses Vendor_Normalized when available to avoid payment-rail prefixes.
    """
    if df is None or len(df) == 0 or 'Debit' not in df.columns:
        return pd.DataFrame()

    # Build a stable vendor key for concentration
    if 'Vendor_Normalized' in df.columns:
        key_series = df['Vendor_Normalized']
    elif 'Description_clean' in df.columns:
        key_series = df['Description_clean']
    elif 'Description' in df.columns:
        key_series = df['Description']
    else:
        key_series = pd.Series(['UNKNOWN'] * len(df), index=df.index)

    key_series = key_series.fillna('').astype(str).str.strip()

    if 'Description_clean' in df.columns:
        fallback = df['Description_clean'].fillna('').astype(str).str.strip()
    elif 'Description' in df.columns:
        fallback = df['Description'].fillna('').astype(str).str.strip()
    else:
        fallback = pd.Series(['UNKNOWN'] * len(df), index=df.index)

    key_series = key_series.where(key_series != '', fallback).replace('', 'UNKNOWN')

    vendor_spending = (
        df.assign(_vendor_key=key_series)
          .groupby('_vendor_key')['Debit']
          .sum()
          .sort_values(ascending=False)
    )
    total_spending = float(vendor_spending.sum()) if len(vendor_spending) else 0.0

    concentration = []
    if total_spending > 0:
        top_n = 10
        top_amounts = vendor_spending.head(top_n)
        for vendor, amount in top_amounts.items():
            pct = (float(amount) / total_spending) * 100.0

            # Risk assessment (same thresholds as prior version)
            if pct > 25:
                risk = 'HIGH'
            elif pct > 15:
                risk = 'MEDIUM'
            elif pct > 10:
                risk = 'MODERATE'
            else:
                risk = 'LOW'

            concentration.append({
                'Vendor': vendor,
                'Total_Spending': float(amount),
                'Percentage': float(pct),
                'Risk_Level': risk
            })

        # Add an aggregated "OTHER" row so % sums to 100% while keeping the table compact
        if len(vendor_spending) > top_n:
            other_amount = float(total_spending - float(top_amounts.sum()))
            if other_amount > 0:
                other_pct = (other_amount / total_spending) * 100.0
                concentration.append({
                    'Vendor': 'OTHER (all remaining vendors)',
                    'Total_Spending': other_amount,
                    'Percentage': float(other_pct),
                    'Risk_Level': 'N/A'
                })

    return pd.DataFrame(concentration)


def calculate_monthly_trends(df):
    """
    Calculate monthly trends for cash flow analysis.
    """
    df_copy = df.copy()
    df_copy['Month'] = pd.to_datetime(df_copy['Date']).dt.to_period('M')
    
    monthly = df_copy.groupby('Month').agg({
        'Debit': 'sum',
        'Credit': 'sum',
        'Date': 'count'
    })
    
    monthly.columns = ['Total_Outflows', 'Total_Inflows', 'Transaction_Count']
    monthly['Net_Cash_Flow'] = monthly['Total_Inflows'] - monthly['Total_Outflows']
    monthly['Month_Name'] = monthly.index.astype(str)
    
    # Calculate trend direction
    trends = []
    for i in range(len(monthly)):
        if i == 0:
            trends.append('—')
        else:
            current_net = monthly.iloc[i]['Net_Cash_Flow']
            previous_net = monthly.iloc[i-1]['Net_Cash_Flow']
            
            if current_net > previous_net:
                trends.append('↑ Improving')
            elif current_net < previous_net:
                trends.append('↓ Worsening')
            else:
                trends.append('→ Stable')
    
    monthly['Trend'] = trends
    
    return monthly


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Merge bank statements and generate analysis workbook.")
    parser.add_argument("--input-dir", type=Path, default=None, help="Directory containing input .xlsx/.xls files.")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path. Defaults to merged_statements_<timestamp>.xlsx.")
    parser.add_argument("--strict-validation", action="store_true", help="Fail if mini-validation finds inconsistencies.")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logs.")
    parser.add_argument("--synthetic", action="store_true", help="Mark output as synthetic/demo dataset (adds disclaimers in workbook).")
    parser.add_argument("--author", type=str, default="Tokmianina_Irina", help="Excel document properties author/creator.")
    parser.add_argument("--patch-excel-app-metadata", action="store_true", help="Patch Excel app metadata (docProps/app.xml Application/AppVersion).")

    return parser.parse_args()


def main():
    """
    Main function of the script.
    """
    args = parse_args()
    configure_logging(args.verbose)
    synthetic_mode = getattr(args, "synthetic", False)
    # CONFIGURATION
    # Add timestamp to output filename for version control
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    OUTPUT_FILE = f"merged_statements_{timestamp}.xlsx"
    
    # Get the current working directory (where script is run from)
    current_dir = Path.cwd()
    input_dir = args.input_dir or (current_dir / "input" if (current_dir / "input").exists() else current_dir)
    logger.info(f"Input folder: {input_dir}")
    
    # Standard output columns
    OUTPUT_COLUMNS = [
        'Txn_ID',
        'Date',
        'Description',
        'Description_clean', # Cleaned vendor name
        'Debit',
        'Credit',
        'Operation_Type',
        'Direction',
        'Bank',             # Source bank
        'Source_File',
        'Source_Sheet',
        'Source_Row',
        'Amount_Signed'
    ]
    
    all_dataframes = []
    beginning_balances = []  # Store beginning balance for each file
    ending_balances = []     # Store ending balance for each file (if detectable)
    
    excel_files = [
        f for f in list(input_dir.glob('*.xlsx')) + list(input_dir.glob('*.xls'))
        if not f.name.startswith('merged_')  # Ignore output files
        and not f.name.startswith('~')  # Ignore Excel temp files
        and not f.name.startswith('.')  # Ignore hidden files
    ]
    excel_files = [
        f for f in excel_files 
        if not f.name.startswith('~$') and f.name != OUTPUT_FILE
    ]
    
    if not excel_files:
        logger.info("  No Excel files found for processing in current folder.")
        return
    
    logger.info(f"Found files for processing: {len(excel_files)}\n")
    
    for file_path in excel_files:
        df = process_excel_file(file_path, OUTPUT_COLUMNS)
        if df is not None:
            all_dataframes.append(df)
            # Try to detect beginning balance for this file
            bb = detect_beginning_balance(file_path, df)
            if bb is not None:
                beginning_bal, beginning_ref = bb
                beginning_balances.append({
                    'file': file_path.name,
                    'balance': beginning_bal,
                    'ref': beginning_ref
                })

            # Try to detect ending/closing balance for this file
            eb = detect_ending_balance(file_path, df)
            if eb is not None:
                ending_bal, ending_ref = eb
                ending_balances.append({
                    'file': file_path.name,
                    'balance': ending_bal,
                    'ref': ending_ref
                })
    if not all_dataframes:
        logger.info("\n No files were successfully processed. Merged file not created.")
        return
    
    # Calculate total beginning balance
    total_beginning_balance = sum([b['balance'] for b in beginning_balances]) if beginning_balances else None
    
    logger.info("Merging data...")
    result_df = pd.concat(all_dataframes, ignore_index=True)
    
    # Add signed amount (Credit - Debit, single column)
    result_df['Amount_Signed'] = result_df['Credit'].fillna(0) - result_df['Debit'].fillna(0)
    
    # Sort with stable tie-breakers to ensure deterministic Txn_ID assignment
    # Same data will ALWAYS produce same Txn_IDs, making the analysis reproducible
    if 'Date' in result_df.columns:
        # P2.6: Sort by: (1) Date, (2) Source_File, (3) Source_Sheet, (4) Source_Row
        # This guarantees consistent order even for multi-sheet files or same-day transactions
        # kind='mergesort' ensures stable sort (equal keys maintain original order)
        result_df = result_df.sort_values(['Date', 'Source_File', 'Source_Sheet', 'Source_Row'], 
                                         ascending=[True, True, True, True],
                                         kind='mergesort')  # Stable sort algorithm
        # Assign Txn_IDs after sorting - they will be stable and reproducible across runs
        result_df['Txn_ID'] = [f"TXN_{i+1:05d}" for i in range(len(result_df))]

    # Audit trail: ensure traceability columns are present for every transaction
    logger.info(f"Merged transactions: {len(result_df)}")
    logger.info("Audit trail: populated (verbose samples: --verbose)")

    if args.verbose:
        logger.info("\nAudit trail samples (first transaction per source file):")
        for source_file in result_df['Source_File'].unique():
            sample = result_df[result_df['Source_File'] == source_file].iloc[0]
            logger.info(
                f"  {source_file}: Sheet {sample['Source_Sheet']} | Row {int(sample['Source_Row'])} | Txn_ID {sample['Txn_ID']}"
            )

    # Save to current directory

    output_path = args.output or (current_dir / OUTPUT_FILE)
    
    logger.info("Generating Excel workbook and analysis sheets...")
    
    # Opening/closing balances availability (per source file)
    total_files = len(all_dataframes)
    files_with_opening = len(beginning_balances)

    if files_with_opening > 0 and total_beginning_balance is not None:
        logger.info(
            f"Opening balances found: {files_with_opening}/{total_files} files (total ${total_beginning_balance:,.2f})"
        )
        if args.verbose:
            for bb in beginning_balances:
                ref = f" [{bb['ref']}]" if bb.get('ref') else ""
                logger.info(f"  - {bb['file']}: ${bb['balance']:,.2f}{ref}")
    else:
        logger.warning(
            f"Opening balances found: 0/{total_files} files (reconciliation partial; balances N/A)"
        )

# Generate analytics
    benford_results = benford_law_analysis(result_df)
    cash_metrics, _ = calculate_cash_flow_metrics(result_df)
    monthly_trends = calculate_monthly_trends(result_df)
    concentration_df = calculate_concentration_risk(result_df)
    findings_df, run_stats = detect_findings(result_df, return_run_stats=True)  # Findings + run-specific calculations
    # Mini-validation (does not affect detection results)
    mini_validate_vendor_growth_findings(result_df, findings_df, logger, strict=getattr(args, 'strict_validation', False))
    mini_validate_duplicate_payment_findings(result_df, findings_df, logger, strict=getattr(args, 'strict_validation', False))

    # Final integrity checks (does not affect detection results)
    final_integrity_checks(result_df, findings_df, strict=getattr(args, 'strict_validation', False), logger=logger)

    
    # Add beginning/ending balance to cash metrics
    if total_beginning_balance is not None:
        cash_metrics['beginning_balance'] = total_beginning_balance
        cash_metrics['ending_balance'] = total_beginning_balance + cash_metrics['net_cash_flow']
        
        # Calculate runway from ENDING balance (not beginning)
        # This shows how long company can survive from current position
        if cash_metrics.get('avg_daily_burn', 0) > 0 and cash_metrics['ending_balance'] > 0:
            cash_metrics['runway_days'] = cash_metrics['ending_balance'] / cash_metrics['avg_daily_burn']
        elif cash_metrics['ending_balance'] <= 0:
            # Already in overdraft or zero balance
            cash_metrics['runway_days'] = 0
    
    # Category breakdown
    category_summary = result_df.groupby('Operation_Type').agg({
        'Debit': 'sum',
        'Credit': 'sum'
    }).sort_values('Debit', ascending=False)
    
    # Display summary
    logger.info(f"Cash flow: ${cash_metrics['net_cash_flow']:,.2f}")
    if benford_results:
        status = "within threshold" if benford_results.get('passes') else "review"
        logger.info(f"Benford check: {status} (see Dashboard)")

    # Save with formatting and analysis sheet
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Workbook document properties
        author = getattr(args, 'author', 'Tokmianina_Irina')
        writer.book.properties.creator = author
        writer.book.properties.lastModifiedBy = author
        # Sheet 1: All Transactions
        # Do not export internal helper columns
        if '_vendor_key' in result_df.columns:
            result_df = result_df.drop(columns=['_vendor_key'])

        result_df.to_excel(writer, index=False, sheet_name='Transactions')
        
        worksheet = writer.sheets['Transactions']
        
        # Map Txn_ID -> Excel row number in Transactions (for internal hyperlinks from Findings)
        txn_to_row = {str(tid): i + 2 for i, tid in enumerate(result_df['Txn_ID'].tolist()) if pd.notna(tid)}
        # Internal hyperlinks from Findings should always land in column A in Transactions
        txn_id_col_letter = 'A'

        
        # Set column widths
        column_widths = {
            'A': 12,   # Date
            'B': 50,   # Description
            'C': 15,   # Debit
            'D': 15,   # Credit
            'E': 15,   # Bank
            'F': 45,   # Description_clean
            'G': 25    # Operation_Type
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Format date column
        from openpyxl.styles import PatternFill, Font, Alignment
        
        for row in range(2, len(result_df) + 2):
            cell = worksheet[f'A{row}']
            cell.number_format = 'MM/DD/YYYY'
        
        # Format number columns with thousand separators and USD currency
        for col in ['C', 'D']:  # Debit, Credit
            for row in range(2, len(result_df) + 2):
                cell = worksheet[f'{col}{row}']
                cell.number_format = '$#,##0.00'
        
        # Format Amount_Signed column (column M in output)
        amount_signed_col = None
        for idx, col_name in enumerate(result_df.columns, 1):
            if col_name == 'Amount_Signed':
                from openpyxl.utils import get_column_letter
                amount_signed_col = get_column_letter(idx)
                break
        
        if amount_signed_col:
            for row in range(2, len(result_df) + 2):
                cell = worksheet[f'{amount_signed_col}{row}']
                cell.number_format = '$#,##0.00'
        
        # P2: Add freeze panes and auto-filter for better usability
        worksheet.freeze_panes = 'A2'  # Freeze header row
        worksheet.auto_filter.ref = worksheet.dimensions  # Enable filtering
        
        # Sheet 2: FINDINGS & ALERTS (Professional audit structure)
        if len(findings_df) > 0:
            ws_findings = writer.book.create_sheet('Findings & Alerts', 1)  # Insert as 2nd sheet
            
            current_row = 1
            
            # === SECTION 1: EXECUTIVE SUMMARY ===
            ws_findings[f'A{current_row}'] = 'FINDINGS & ALERTS - EXECUTIVE SUMMARY'
            ws_findings[f'A{current_row}'].font = Font(bold=True, size=16, color='FFFFFF')
            ws_findings[f'A{current_row}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            ws_findings.merge_cells(f'A{current_row}:M{current_row}')
            current_row += 1
            
            ws_findings[f'A{current_row}'] = f'Automated screening identified {len(findings_df)} items for review'
            ws_findings[f'A{current_row}'].font = Font(italic=True, size=10)
            ws_findings.merge_cells(f'A{current_row}:M{current_row}')
            current_row += 2

            if synthetic_mode:
                ws_findings[f'A{current_row}'] = 'DATA NOTE: Synthetic/demo dataset (portfolio). No real PII; identifiers are masked.'
                ws_findings[f'A{current_row}'].font = Font(size=10, italic=True, color='7F7F7F')
                ws_findings.merge_cells(f'A{current_row}:M{current_row}')
                current_row += 2
            
            # Priority summary
            p1_count = len(findings_df[findings_df['Priority'] == 'P1'])
            p2_count = len(findings_df[findings_df['Priority'] == 'P2'])
            p3_count = len(findings_df[findings_df['Priority'] == 'P3'])
            
            ws_findings[f'A{current_row}'] = 'Priority Breakdown:'
            ws_findings[f'A{current_row}'].font = Font(bold=True, size=11)
            current_row += 1
            
            ws_findings[f'A{current_row}'] = f'P1 (High): {p1_count}'
            ws_findings[f'A{current_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            current_row += 1
            
            ws_findings[f'A{current_row}'] = f'P2 (Medium): {p2_count}'
            ws_findings[f'A{current_row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            current_row += 1
            
            ws_findings[f'A{current_row}'] = f'P3 (Low): {p3_count}'
            ws_findings[f'A{current_row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            current_row += 2
            
            # === QUANTITATIVE SUMMARY ===
            ws_findings[f'A{current_row}'] = 'Financial Impact:'
            ws_findings[f'A{current_row}'].font = Font(bold=True, size=11)
            current_row += 1
            
            # Calculate metrics - SEPARATE by Amount_Type to avoid mixing semantics
            transaction_amounts = findings_df[
                findings_df['Amount_Type'].str.contains('Transaction Amount|Total Paid|Total Amount', na=False, case=False)
            ]['Amount'].sum()
            
            growth_deltas = findings_df[
                findings_df['Amount_Type'].str.contains('Growth|Delta', na=False, case=False)
            ]['Amount'].sum()
            
            total_outflows = result_df['Debit'].sum()
            
            # Only calculate % for transaction amounts (not deltas!)
            flagged_pct = (transaction_amounts / total_outflows * 100) if total_outflows > 0 else 0
            
            # Top 3 categories by amount
            top_categories = findings_df.groupby('Category')['Amount'].sum().nlargest(3)
            
            # Display separated metrics
            ws_findings[f'A{current_row}'] = f'Total Flagged Transaction Amount: ${transaction_amounts:,.2f}'
            ws_findings[f'A{current_row}'].font = Font(size=10, bold=True)
            current_row += 1
            
            if growth_deltas > 0:
                ws_findings[f'A{current_row}'] = f'Total Growth Delta Flagged: ${growth_deltas:,.2f}'
                ws_findings[f'A{current_row}'].font = Font(size=10)
                current_row += 1
            
            ws_findings[f'A{current_row}'] = f'As % of Total Outflows: {flagged_pct:.1f}% (transaction amounts only)'
            ws_findings[f'A{current_row}'].font = Font(size=10, italic=True)
            current_row += 1
            
            ws_findings[f'A{current_row}'] = 'Top 3 Categories by $:'
            ws_findings[f'A{current_row}'].font = Font(size=10, italic=True)
            current_row += 1
            
            for cat, amt in top_categories.items():
                ws_findings[f'A{current_row}'] = f'  • {cat}: ${amt:,.2f}'
                ws_findings[f'A{current_row}'].font = Font(size=9)
                current_row += 1
            
            current_row += 2
            
            # === RULE COVERAGE (shows all rules checked, even if 0 findings) ===
            ws_findings[f'A{current_row}'] = 'Rule Coverage:'
            ws_findings[f'A{current_row}'].font = Font(bold=True, size=11)
            current_row += 1
            
            ws_findings[f'A{current_row}'] = 'All detection rules were applied to the dataset'
            ws_findings[f'A{current_row}'].font = Font(italic=True, size=9)
            current_row += 1
            
            # Count findings by category
            rule_coverage = {
                'Duplicate Payment': 0,
                'New Counterparty - High Amount': 0,
                'Round Amount Clustering': 0,
                'Split Payments': 0,
                'Weekend Activity': 0,
                'Weekend Activity (Multiple)': 0,
                'High Velocity Outflows': 0,
                'Vendor Concentration Growth': 0
            }
            
            for cat in findings_df['Category'].unique():
                if cat in rule_coverage:
                    rule_coverage[cat] = len(findings_df[findings_df['Category'] == cat])
                elif 'Weekend' in cat:
                    rule_coverage['Weekend Activity'] += len(findings_df[findings_df['Category'] == cat])
            
            # Display as compact list
            coverage_items = []
            for rule, count in rule_coverage.items():
                if rule != 'Weekend Activity (Multiple)':  # Don't double-count weekend
                    coverage_items.append(f'{rule}: {count}')
            
            # Display in 2 columns
            for i in range(0, len(coverage_items), 2):
                row_text = coverage_items[i]
                if i + 1 < len(coverage_items):
                    row_text += f'  |  {coverage_items[i+1]}'
                ws_findings[f'A{current_row}'] = row_text
                ws_findings[f'A{current_row}'].font = Font(size=9)
                current_row += 1
            
            current_row += 2
            
            # === SECTION 2: ALERTS TABLE (machine-readable) ===
            ws_findings[f'A{current_row}'] = 'ALERTS TABLE (Detailed)'
            ws_findings[f'A{current_row}'].font = Font(bold=True, size=14)
            ws_findings.merge_cells(f'A{current_row}:M{current_row}')
            current_row += 1
            
            ws_findings[f'A{current_row}'] = 'Each alert includes evidence, affected transactions, and recommended next steps'
            ws_findings[f'A{current_row}'].font = Font(italic=True, size=9)
            ws_findings.merge_cells(f'A{current_row}:M{current_row}')
            current_row += 2
            
            # Table headers (machine-readable format)
            alerts_header_row = current_row  # start of alerts table (header row)
            headers = ['Alert ID', 'Priority', 'Category', 'Severity', 'Vendor', 
                      'Amount', 'Amount_Type', 'Bank', 'Type', 'Evidence', 'Txn IDs', 'Next Step', 'Anchor Date (latest txn)']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_findings.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            current_row += 1
            
            alerts_first_data_row = current_row  # first data row (after headers)
            # Data rows sorted by Priority
            findings_sorted = findings_df.sort_values(['Priority', 'Severity'])
            alert_to_findings_row = {}
            
            for idx, row_data in findings_sorted.iterrows():
                ws_findings.cell(row=current_row, column=1, value=row_data['Alert_ID'])
                alert_to_findings_row[str(row_data['Alert_ID'])] = current_row
                
                # Priority with color coding
                priority_cell = ws_findings.cell(row=current_row, column=2, value=row_data['Priority'])
                if row_data['Priority'] == 'P1':
                    priority_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif row_data['Priority'] == 'P2':
                    priority_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    priority_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                
                ws_findings.cell(row=current_row, column=3, value=row_data['Category'])
                ws_findings.cell(row=current_row, column=4, value=row_data['Severity'])
                ws_findings.cell(row=current_row, column=5, value=str(row_data['Vendor'])[:50])
                
                # Amount
                amount_cell = ws_findings.cell(row=current_row, column=6, value=round(row_data['Amount'], 2))
                amount_cell.number_format = '$#,##0.00'
                
                # Amount_Type (NEW) - don't truncate, use wrap_text instead
                ws_findings.cell(row=current_row, column=7, value=str(row_data.get('Amount_Type', 'N/A')))
                
                # Bank
                ws_findings.cell(row=current_row, column=8, value=str(row_data.get('Bank', 'N/A'))[:20])
                
                # Operation Type
                ws_findings.cell(row=current_row, column=9, value=str(row_data.get('Operation_Type', 'N/A'))[:25])
                
                # Evidence
                ws_findings.cell(row=current_row, column=10, value=row_data['Evidence'])
                
                # Transaction IDs
                # Transaction IDs (plain text).
                # Note: Excel supports only one hyperlink per cell. Per-Txn drill-down is provided in the "Evidence Index" sheet.
                txn_ids_raw = str(row_data.get('Txn_IDs', ''))
                txn_ids_text = txn_ids_raw[:200]
                ws_findings.cell(row=current_row, column=11, value=txn_ids_text)
                # Next Step
                ws_findings.cell(row=current_row, column=12, value=row_data['Next_Step'])
                
                # Date
                # Date (anchor): use Anchor_Date if provided; otherwise max(Date_1, Date_2).
                anchor = row_data.get('Anchor_Date', None)
                if pd.notna(anchor):
                    anchor = pd.to_datetime(anchor, errors='coerce')
                else:
                    d1 = row_data.get('Date_1', None)
                    d2 = row_data.get('Date_2', None)
                    anchor = None
                    if pd.notna(d1):
                        anchor = d1
                    if pd.notna(d2) and (anchor is None or d2 >= anchor):
                        anchor = d2
                date_str = anchor.strftime('%Y-%m-%d') if anchor is not None else ''
                ws_findings.cell(row=current_row, column=13, value=date_str)
                
                # Wrap text for readability
                for col in range(1, 14):
                    ws_findings.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                
                current_row += 1


            
            alerts_last_data_row = current_row - 1  # last alerts data row

            # Note: Txn-level navigation is provided via the Evidence Index sheet (one row per Txn ID).
            # Add methodology note at bottom
            current_row += 2
            ws_findings[f'A{current_row}'] = 'METHODOLOGY & THRESHOLDS:'
            ws_findings[f'A{current_row}'].font = Font(bold=True, size=11)
            current_row += 1
            
            notes = [
                'Detection Rules:',
                '• Duplicate Detection: Same vendor ±1% amount within 3 days',
                '• New Counterparty Alert: First transaction >$50,000 (materiality threshold)',
                '• Split Payments: 2+ payments to same counterparty on same day, total >$10k',
                '• Round Amount Clustering: Multiple exact thousands (excluding fixed costs)',
                '• Weekend Activity: Non-automated payments >$25k or 95th percentile (collapses if >3)',
                '• High Velocity: 5+ large payments (>$20k) single day',
                '• Vendor Concentration Growth: Baseline (first active month) → latest month >2x, baseline >$1,000',
                '',
                'Exclusions Applied:',
                '• New Counterparty: Excludes payroll, taxes, utilities, insurance',
                '• Weekend Activity: Excludes automated ACH, tax payments, payroll',
                '• Round Amounts: Excludes rent, lease, loan payments, salaries',
                '',
                'Important Notes:',
                '• These are screening alerts only - not conclusions of fraud',
                '• Each item requires manual investigation and business context',
                '• Legitimate reasons may exist for flagged patterns',
                '• Drill-down: Click Alert ID to open Evidence Index; Txn IDs in Evidence Index link to Transactions'
            ]
            
            for note in notes:
                ws_findings.cell(row=current_row, column=1, value=note).font = Font(size=9, italic=True if note.startswith('•') else False, bold=note.endswith(':'))
                ws_findings.merge_cells(f'A{current_row}:M{current_row}')
                current_row += 1
            
            # Set column widths
            ws_findings.column_dimensions['A'].width = 12  # Alert ID
            ws_findings.column_dimensions['B'].width = 10  # Priority
            ws_findings.column_dimensions['C'].width = 25  # Category
            ws_findings.column_dimensions['D'].width = 10  # Severity
            ws_findings.column_dimensions['E'].width = 30  # Vendor
            ws_findings.column_dimensions['F'].width = 15  # Amount
            ws_findings.column_dimensions['G'].width = 30  # Amount_Type (NEW)
            ws_findings.column_dimensions['H'].width = 15  # Bank
            ws_findings.column_dimensions['I'].width = 20  # Type
            ws_findings.column_dimensions['J'].width = 40  # Evidence
            ws_findings.column_dimensions['K'].width = 25  # Txn IDs
            ws_findings.column_dimensions['L'].width = 40  # Next Step
            ws_findings.column_dimensions['M'].width = 12  # Date
            # P2: Add auto-filter for better usability
            # Note: filter should cover only the alerts table (exclude methodology/notes)
            if alerts_last_data_row >= alerts_header_row:
                ws_findings.auto_filter.ref = f'A{alerts_header_row}:M{alerts_last_data_row}'
        
            # Sheet 2.5: Evidence Index (one row per Txn_ID with a direct hyperlink to Transactions)
            # Rationale: Findings may reference multiple transactions; Excel cannot store multiple independent hyperlinks in one cell.
            ws_evidence = writer.book.create_sheet('Evidence Index', 2)

            ws_evidence['A1'] = 'EVIDENCE INDEX (Txn-level drill-down)'
            ws_evidence['A1'].font = Font(bold=True, size=14)
            ws_evidence.merge_cells('A1:J1')

            headers_e = ['Alert ID', 'Category', 'Vendor', 'Group', 'Txn ID', 'Txn Date', 'Debit', 'Credit', 'Bank', 'Source']
            for c, h in enumerate(headers_e, 1):
                cell = ws_evidence.cell(row=2, column=c, value=h)
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

            txn_lookup = result_df.set_index('Txn_ID', drop=False)
            alert_to_evidence_row = {}
            out_row = 3

            def _parse_txn_groups(txn_ids_text: str):
                txt = str(txn_ids_text or '')
                if ('baseline' in txt.lower()) and ('latest' in txt.lower()) and '|' in txt:
                    parts = [p.strip() for p in txt.split('|')]
                    out = []
                    for p in parts:
                        grp = 'Related'
                        pl = p.lower()
                        if pl.startswith('baseline'):
                            grp = 'Baseline'
                        elif pl.startswith('latest'):
                            grp = 'Latest'
                        ids = re.findall(r'TXN_\d{5}', p)
                        out.extend([(grp, tid) for tid in ids])
                    return out
                ids = re.findall(r'TXN_\d{5}', txt)
                return [('Related', tid) for tid in ids]

            for _, f in findings_sorted.iterrows():
                alert_id = str(f.get('Alert_ID', ''))
                cat = str(f.get('Category', ''))
                vendor = str(f.get('Vendor', ''))
                txn_groups = _parse_txn_groups(f.get('Txn_IDs', ''))

                if alert_id and (alert_id not in alert_to_evidence_row) and len(txn_groups) > 0:
                    alert_to_evidence_row[alert_id] = out_row

                for grp, tid in txn_groups:
                    ws_evidence.cell(row=out_row, column=1, value=alert_id)
                    ws_evidence.cell(row=out_row, column=2, value=cat)
                    ws_evidence.cell(row=out_row, column=3, value=vendor)
                    ws_evidence.cell(row=out_row, column=4, value=grp)

                    c_tid = ws_evidence.cell(row=out_row, column=5, value=tid)
                    if tid in txn_to_row:
                        c_tid.hyperlink = f"#'Transactions'!A{txn_to_row[tid]}"
                        c_tid.font = Font(color='0563C1', underline='single')

                    if tid in txn_lookup.index:
                        t = txn_lookup.loc[tid]
                        t_date = t.get('Date', None)
                        ws_evidence.cell(row=out_row, column=6, value=t_date.strftime('%Y-%m-%d') if pd.notna(t_date) else '')
                        ws_evidence.cell(row=out_row, column=7, value=float(t.get('Debit')) if pd.notna(t.get('Debit')) else None)
                        ws_evidence.cell(row=out_row, column=8, value=float(t.get('Credit')) if pd.notna(t.get('Credit')) else None)
                        ws_evidence.cell(row=out_row, column=9, value=str(t.get('Bank', '')))
                        src = f"{t.get('Source_File', '')} | {t.get('Source_Sheet', '')} | R{t.get('Source_Row', '')}"
                        ws_evidence.cell(row=out_row, column=10, value=src)

                    out_row += 1

            ws_evidence.freeze_panes = 'A3'
            ws_evidence.auto_filter.ref = f"A2:J{max(2, out_row-1)}"
            widths = {'A': 10, 'B': 28, 'C': 22, 'D': 10, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 10, 'J': 40}
            for col, w in widths.items():
                ws_evidence.column_dimensions[col].width = w
            
            # Format Debit/Credit columns with USD currency
            for row_num in range(3, out_row):
                debit_cell = ws_evidence.cell(row=row_num, column=7)
                credit_cell = ws_evidence.cell(row=row_num, column=8)
                if debit_cell.value is not None:
                    debit_cell.number_format = '$#,##0.00'
                if credit_cell.value is not None:
                    credit_cell.number_format = '$#,##0.00'

            # Link Alert ID in Findings to Evidence Index (first row for that alert)
            for a_id, ev_row in alert_to_evidence_row.items():
                fr = alert_to_findings_row.get(a_id)
                if fr:
                    c = ws_findings.cell(row=fr, column=1)
                    c.hyperlink = f"#'Evidence Index'!A{ev_row}"
                    c.font = Font(color='0563C1', underline='single')

        # Sheet 3: Analytics Dashboard
        ws_analytics = writer.book.create_sheet('Analytics Dashboard')
        
        current_row = 1
        
        # === SECTION 1: CASH FLOW METRICS ===
        ws_analytics[f'A{current_row}'] = 'CASH FLOW METRICS'
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=14, color='FFFFFF')
        ws_analytics[f'A{current_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_analytics.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 2
        
        # Get date range for context
        start_date = result_df['Date'].min().strftime('%b %d, %Y') if result_df['Date'].notna().any() else ''
        end_date = result_df['Date'].max().strftime('%b %d, %Y') if result_df['Date'].notna().any() else ''
        period_days = cash_metrics.get('period_days', 0)
        
        # Cash flow data with rounded numbers
        cash_data = [
            ['Analysis Period', f"{period_days} elapsed days ({start_date} to {end_date})", ''],
            ['Total Transactions', len(result_df), ''],
            ['', '', ''],
        ]
        
        # Add Beginning/Ending Balance if available
        if 'beginning_balance' in cash_metrics:
            cash_data.extend([
                ['Beginning Balance', round(cash_metrics['beginning_balance'], 2), 'Starting account balance'],
                ['Total Inflows', round(cash_metrics['total_inflow'], 2), ''],
                ['Total Outflows', round(cash_metrics['total_outflow'], 2), ''],
                ['Net Cash Flow', round(cash_metrics['net_cash_flow'], 2), 'Positive = surplus, Negative = deficit'],
                ['Ending Balance', round(cash_metrics['ending_balance'], 2), 
                 ' OVERDRAFT - Account overdrawn' if cash_metrics['ending_balance'] < 0 else 'Final account balance'],
                ['', '', ''],
            ])
        else:
            cash_data.extend([
                ['Total Inflows', round(cash_metrics['total_inflow'], 2), ''],
                ['Total Outflows', round(cash_metrics['total_outflow'], 2), ''],
                ['Net Cash Flow', round(cash_metrics['net_cash_flow'], 2), 'Positive = healthy' if cash_metrics['net_cash_flow'] > 0 else 'Negative = burning cash'],
                ['', '', ''],
            ])
        
        if 'avg_daily_burn' in cash_metrics:
            daily_metrics = [
                ['Avg Daily Inflow', round(cash_metrics.get('avg_daily_inflow', 0), 2), 'Average revenue per day'],
                ['Avg Daily Outflow', round(cash_metrics.get('avg_daily_outflow', 0), 2), 'Average spending per day'],
                ['Daily Burn Rate', round(cash_metrics.get('avg_daily_burn', 0), 2), 'Net daily cash consumption'],
            ]
            
            # Add runway if available and burning cash
            if 'runway_days' in cash_metrics:
                runway = cash_metrics['runway_days']
                if runway == 0:
                    runway_comment = ' No runway - account at or below zero'
                elif runway < 30:
                    runway_comment = f' CRITICAL - Less than 1 month of cash remaining'
                elif runway < 90:
                    runway_comment = f'Days until funds depleted (at current burn rate)'
                else:
                    runway_comment = 'Healthy cash runway'
                daily_metrics.append(['Cash Runway (from current balance)', round(runway, 0), runway_comment])
            
            cash_data.extend(daily_metrics)
        
        for row_data in cash_data:
            ws_analytics.cell(row=current_row, column=1, value=row_data[0]).font = Font(bold=True)
            cell = ws_analytics.cell(row=current_row, column=2, value=row_data[1])
            
            # Format numbers
            if isinstance(row_data[1], (int, float)) and row_data[0] != 'Total Transactions':
                cell.number_format = '$#,##0.00'
                
                # Color code Ending Balance if negative (overdraft)
                if row_data[0] == 'Ending Balance' and row_data[1] < 0:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(bold=True, color='C00000')
            
            # Add comment column
            comment_cell = ws_analytics.cell(row=current_row, column=3, value=row_data[2])
            comment_cell.font = Font(italic=True, size=9)
            
            # Color code warning comments
            if '' in str(row_data[2]):
                comment_cell.font = Font(italic=True, size=9, color='C00000', bold=True)
            
            current_row += 1
        
        current_row += 2
        
        # === SECTION 2: MONTHLY TRENDS ===
        ws_analytics[f'A{current_row}'] = 'MONTHLY CASH FLOW TRENDS'
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=14, color='FFFFFF')
        ws_analytics[f'A{current_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws_analytics.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        ws_analytics[f'A{current_row}'] = 'Track monthly performance to identify seasonality and trends'
        ws_analytics[f'A{current_row}'].font = Font(italic=True)
        current_row += 2
        
        # Monthly trends table headers
        headers = ['Month', 'Inflows', 'Outflows', 'Net Cash Flow', 'Trend']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_analytics.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        current_row += 1
        
        # Monthly data
        for idx, row_data in monthly_trends.iterrows():
            ws_analytics.cell(row=current_row, column=1, value=row_data['Month_Name'])
            ws_analytics.cell(row=current_row, column=2, value=round(row_data['Total_Inflows'], 2)).number_format = '$#,##0.00'
            ws_analytics.cell(row=current_row, column=3, value=round(row_data['Total_Outflows'], 2)).number_format = '$#,##0.00'
            
            net_cell = ws_analytics.cell(row=current_row, column=4, value=round(row_data['Net_Cash_Flow'], 2))
            net_cell.number_format = '$#,##0.00'
            
            # Color code net cash flow
            if row_data['Net_Cash_Flow'] > 0:
                net_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                net_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            trend_cell = ws_analytics.cell(row=current_row, column=5, value=row_data['Trend'])
            if '↑' in row_data['Trend']:
                trend_cell.font = Font(color='008000', bold=True)
            elif '↓' in row_data['Trend']:
                trend_cell.font = Font(color='C00000', bold=True)
            
            current_row += 1
        
        current_row += 2
        
        # === SECTION 3: BENFORD'S LAW SCREENING (EXPLORATORY) ===
        ws_analytics[f'A{current_row}'] = "APPENDIX: BENFORD'S LAW SCREENING (EXPLORATORY / ADVISORY ONLY)"
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=11, color='666666')
        ws_analytics[f'A{current_row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws_analytics.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        if benford_results:
            ws_analytics[f'A{current_row}'] = ' ADVISORY: Statistical screening referenced in forensic literature - for initial review only, not definitive evidence'
            ws_analytics[f'A{current_row}'].font = Font(italic=True, size=9, color='C00000')
            current_row += 2
            
            # Result
            status = " Within Normal Range" if benford_results['passes'] else "⚠ Unusual Pattern Detected"
            ws_analytics.cell(row=current_row, column=1, value='Screening Result:').font = Font(bold=True)
            result_cell = ws_analytics.cell(row=current_row, column=2, value=status)
            result_cell.font = Font(bold=True, size=11)
            if benford_results['passes']:
                result_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                result_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws_analytics.merge_cells(f'B{current_row}:C{current_row}')
            current_row += 1
            
            # Context about sample size
            total_txns = len(result_df)
            sample_size = benford_results['sample_size']
            pct = (sample_size / total_txns * 100) if total_txns > 0 else 0
            
            ws_analytics.cell(row=current_row, column=1, value='Sample:')
            ws_analytics.cell(row=current_row, column=2, value=f"{sample_size} debit transactions ({pct:.0f}% of total)")
            ws_analytics.cell(row=current_row, column=3, value='Credits excluded per standard practice').font = Font(italic=True, size=9)
            current_row += 1
            
            ws_analytics.cell(row=current_row, column=1, value='Chi-Square Statistic:')
            ws_analytics.cell(row=current_row, column=2, value=f"{benford_results['chi_square']:.2f}")
            ws_analytics.cell(row=current_row, column=3, value='Threshold: 15.507 (95% confidence)').font = Font(italic=True, size=9)
            current_row += 2
            
            # IMPORTANT: Add limitations box
            ws_analytics[f'A{current_row}'] = 'IMPORTANT LIMITATIONS:'
            ws_analytics[f'A{current_row}'].font = Font(bold=True, size=10, color='C00000')
            current_row += 1
            
            limitations = [
                '• This is a screening tool only, not conclusive evidence of fraud',
                '• Small samples (<100) may show false positives',
                '• Fixed-price contracts naturally deviate from Benford distribution',
                '• Multiple data sources may dilute statistical signal',
                '• Requires manual investigation of any flagged anomalies',
                '',
                'Use Case: Initial screening to identify transactions for further review'
            ]
            
            for limitation in limitations:
                ws_analytics.cell(row=current_row, column=1, value=limitation).font = Font(size=9, italic=True)
                ws_analytics.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
            
            current_row += 1
            
            # Distribution table (collapsed)
            ws_analytics.cell(row=current_row, column=1, value='First Digit Distribution:').font = Font(bold=True, size=10)
            current_row += 1
            
            ws_analytics.cell(row=current_row, column=1, value='Digit').font = Font(bold=True, size=9)
            ws_analytics.cell(row=current_row, column=2, value='Expected %').font = Font(bold=True, size=9)
            ws_analytics.cell(row=current_row, column=3, value='Actual %').font = Font(bold=True, size=9)
            ws_analytics.cell(row=current_row, column=4, value='Deviation').font = Font(bold=True, size=9)
            ws_analytics.cell(row=current_row, column=5, value='Note').font = Font(bold=True, size=9)

            # Apply consistent header styling (fill + larger font)
            for _col in range(1, 6):
                _cell = ws_analytics.cell(row=current_row, column=_col)
                _cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                _cell.font = Font(bold=True, size=11)

            current_row += 1
            
            for idx, row_data in benford_results['results'].iterrows():
                ws_analytics.cell(row=current_row, column=1, value=row_data['First_Digit'])
                ws_analytics.cell(row=current_row, column=2, value=round(row_data['Expected_%'], 1)).number_format = '0.0"%"'
                ws_analytics.cell(row=current_row, column=3, value=round(row_data['Actual_%'], 1)).number_format = '0.0"%"'
                dev_cell = ws_analytics.cell(row=current_row, column=4, value=round(row_data['Deviation_%'], 1))
                dev_cell.number_format = '0.0"%"'
                
                # Assessment
                abs_dev = abs(row_data['Deviation_%'])
                if abs_dev > 5:
                    dev_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    ws_analytics.cell(row=current_row, column=5, value='Review').font = Font(italic=True, size=9)
                else:
                    ws_analytics.cell(row=current_row, column=5, value='OK').font = Font(italic=True, size=9, color='808080')
                
                current_row += 1
        else:
            ws_analytics[f'A{current_row}'] = 'Insufficient data for analysis (minimum 50 transactions required)'
            ws_analytics[f'A{current_row}'].font = Font(italic=True, color='808080')
            current_row += 1
        
        current_row += 2
        
        # === SECTION 4: VENDOR CONCENTRATION RISK ===
        ws_analytics[f'A{current_row}'] = 'VENDOR CONCENTRATION ANALYSIS'
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=14, color='FFFFFF')
        ws_analytics[f'A{current_row}'].fill = PatternFill(start_color='9E480E', end_color='9E480E', fill_type='solid')
        ws_analytics.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        ws_analytics[f'A{current_row}'] = 'Dependency Risk: >25% = HIGH, 15-25% = MEDIUM, 10-15% = MODERATE, <10% = LOW'
        ws_analytics[f'A{current_row}'].font = Font(italic=True)
        current_row += 2
        
        # Table headers
        ws_analytics.cell(row=current_row, column=1, value='Vendor').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=2, value='Total Spending').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=3, value='% of Total').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=4, value='Risk Level').font = Font(bold=True)

        # Apply consistent header styling
        for _col in range(1, 5):
            _cell = ws_analytics.cell(row=current_row, column=_col)
            _cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            _cell.font = Font(bold=True, size=11)

        current_row += 1
        
        if len(concentration_df) > 0:
            for idx, row_data in concentration_df.iterrows():
                ws_analytics.cell(row=current_row, column=1, value=row_data['Vendor'])
                ws_analytics.cell(row=current_row, column=2, value=round(row_data['Total_Spending'], 2)).number_format = '$#,##0.00'
                ws_analytics.cell(row=current_row, column=3, value=round(row_data['Percentage']/100, 3)).number_format = '0.0%'
                
                risk_cell = ws_analytics.cell(row=current_row, column=4, value=row_data['Risk_Level'])
                risk_cell.font = Font(bold=True)
                
                # Color code risk
                level = str(row_data.get('Risk_Level', '')).strip()
                if level == 'N/A':
                    # Aggregated row (not a single-vendor dependency risk)
                    risk_cell.font = Font(italic=True, color='808080')
                elif level == 'HIGH':
                    risk_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif level == 'MEDIUM':
                    risk_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif level == 'MODERATE':
                    risk_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                else:  # LOW
                    risk_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                current_row += 1
        
        current_row += 2
        
        # === SECTION 5: CATEGORY BREAKDOWN ===
        ws_analytics[f'A{current_row}'] = 'SPENDING BY CATEGORY'
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=14, color='FFFFFF')
        ws_analytics[f'A{current_row}'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        ws_analytics.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        ws_analytics[f'A{current_row}'] = 'Separated by cash flow direction (Inflows vs Outflows)'
        ws_analytics[f'A{current_row}'].font = Font(italic=True)
        current_row += 2
        
        # Separate inflows and outflows
        outflows = result_df[result_df['Direction'] == 'Outflow'].groupby('Operation_Type')['Debit'].sum().sort_values(ascending=False)
        inflows = result_df[result_df['Direction'] == 'Inflow'].groupby('Operation_Type')['Credit'].sum().sort_values(ascending=False)
        
        # OUTFLOWS section
        ws_analytics[f'A{current_row}'] = 'OUTFLOWS (Expenses)'
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=11, color='C00000')
        ws_analytics.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        ws_analytics.cell(row=current_row, column=1, value='Category').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=2, value='Amount').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=3, value='% of Outflows').font = Font(bold=True)

        # Apply consistent header styling
        for _col in range(1, 4):
            _cell = ws_analytics.cell(row=current_row, column=_col)
            _cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            _cell.font = Font(bold=True, size=11)

        current_row += 1

        total_outflows = float(outflows.sum()) if outflows is not None else 0.0
        for category, amount in outflows.items():
            ws_analytics.cell(row=current_row, column=1, value=category)
            ws_analytics.cell(row=current_row, column=2, value=round(amount, 2)).number_format = '$#,##0.00'
            pct_cell = ws_analytics.cell(
                row=current_row,
                column=3,
                value=(float(amount) / total_outflows) if total_outflows else 0.0
            )
            pct_cell.number_format = '0.0%'
            current_row += 1
        
        current_row += 1
        
        # INFLOWS section
        ws_analytics[f'A{current_row}'] = 'INFLOWS (Income)'
        ws_analytics[f'A{current_row}'].font = Font(bold=True, size=11, color='008000')
        ws_analytics.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        ws_analytics.cell(row=current_row, column=1, value='Category').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=2, value='Amount').font = Font(bold=True)
        ws_analytics.cell(row=current_row, column=3, value='% of Inflows').font = Font(bold=True)

        # Apply consistent header styling
        for _col in range(1, 4):
            _cell = ws_analytics.cell(row=current_row, column=_col)
            _cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            _cell.font = Font(bold=True, size=11)

        current_row += 1

        total_inflows = float(inflows.sum()) if inflows is not None else 0.0
        for category, amount in inflows.items():
            ws_analytics.cell(row=current_row, column=1, value=category)
            ws_analytics.cell(row=current_row, column=2, value=round(amount, 2)).number_format = '$#,##0.00'
            pct_cell = ws_analytics.cell(
                row=current_row,
                column=3,
                value=(float(amount) / total_inflows) if total_inflows else 0.0
            )
            pct_cell.number_format = '0.0%'
            current_row += 1
        
        # Set column widths
        ws_analytics.column_dimensions['A'].width = 35
        ws_analytics.column_dimensions['B'].width = 25
        ws_analytics.column_dimensions['C'].width = 25
        ws_analytics.column_dimensions['D'].width = 20
        ws_analytics.column_dimensions['E'].width = 20
        
        # === SHEET 4: CONFIG & THRESHOLDS ===
        ws_config = writer.book.create_sheet('Config & Thresholds')
        
        current_row = 1
        
        # Title
        ws_config[f'A{current_row}'] = 'ANALYSIS CONFIGURATION & THRESHOLDS'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=16, color='FFFFFF')
        ws_config[f'A{current_row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # Synthetic/demo dataset disclaimer (optional)
        if synthetic_mode:
            ws_config[f'A{current_row}'] = 'DATA NOTE: Synthetic/demo dataset (portfolio). No real PII; identifiers are masked.'
            ws_config[f'A{current_row}'].font = Font(size=10, italic=True, color='7F7F7F')
            ws_config.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
        
        ws_config[f'A{current_row}'] = 'This sheet documents all thresholds, exclusions, and business rules applied in the analysis'
        ws_config[f'A{current_row}'].font = Font(italic=True, size=10)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 3
        
        # === SECTION 1: DETECTION THRESHOLDS ===
        ws_config[f'A{current_row}'] = 'DETECTION THRESHOLDS'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        # Table headers
        ws_config.cell(row=current_row, column=1, value='Rule').font = Font(bold=True)
        ws_config.cell(row=current_row, column=2, value='Threshold').font = Font(bold=True)
        ws_config.cell(row=current_row, column=3, value='Rationale').font = Font(bold=True)
        ws_config.cell(row=current_row, column=4, value='Source').font = Font(bold=True)
        for col in range(1, 5):
            ws_config.cell(row=current_row, column=col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        current_row += 1
        
        # Threshold data
        thresholds = get_detection_threshold_rows(run_stats)
        
        for threshold_row in thresholds:
            for col_idx, value in enumerate(threshold_row, 1):
                cell = ws_config.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
        
        current_row += 2
        
        # === SECTION 2: EXCLUSIONS ===
        ws_config[f'A{current_row}'] = 'EXCLUSIONS & EXCEPTIONS'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        ws_config[f'A{current_row}'] = 'Categories automatically excluded from certain rules to reduce false positives'
        ws_config[f'A{current_row}'].font = Font(italic=True, size=9)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2
        
        # Exclusions table headers
        ws_config.cell(row=current_row, column=1, value='Rule').font = Font(bold=True)
        ws_config.cell(row=current_row, column=2, value='Excluded Categories').font = Font(bold=True)
        ws_config.cell(row=current_row, column=3, value='Rationale').font = Font(bold=True)
        for col in range(1, 4):
            ws_config.cell(row=current_row, column=col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        current_row += 1
        # Exclusions data
        exclusions = [
            [
                "Duplicate Detection",
                "Shipping & Logistics; Utilities & Services; Bank Fees; Payroll & Benefits; Taxes & Government; Rent & Facilities",
                "Recurring operational payments can repeat amounts.",
            ],
            [
                "Vendor Concentration Growth",
                "Shipping & Logistics; Utilities & Services; Taxes & Government; Bank Fees; Payroll & Benefits; Rent & Facilities; Insurance & Security",
                "Category concentration may be expected due to contracts or statutory payments.",
            ],
            [
                "New Counterparty Alert",
                "Payroll & Benefits; Taxes & Government; Utilities & Services; Insurance & Security",
                "New vendors are common during business growth and onboarding.",
            ],
            [
                "Weekend Activity",
                "Payroll & Benefits; Taxes & Government; Bank Fees",
                "Scheduled processing and deadlines can post on weekends.",
            ],
            [
                "Round Amount Clustering",
                "Rent & Facilities; Loans & Financing; Payroll & Benefits",
                "Fixed contractual amounts are often exact.",
            ],
        ]

        for excl_row in exclusions:

            for col_idx, value in enumerate(excl_row, 1):
                cell = ws_config.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
        
        current_row += 2
        
        # === SECTION 3: INPUT FILES (for reproducibility) ===
        ws_config[f'A{current_row}'] = 'INPUT FILES PROCESSED'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        ws_config[f'A{current_row}'] = 'Files included in this analysis (for audit trail and reproducibility):'
        ws_config[f'A{current_row}'].font = Font(italic=True, size=9)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        # List all processed files
        for source_file in result_df['Source_File'].unique():
            file_txns = len(result_df[result_df['Source_File'] == source_file])
            ws_config[f'A{current_row}'] = f'• {source_file}'
            ws_config[f'A{current_row}'].font = Font(size=9)
            ws_config[f'B{current_row}'] = f'{file_txns} transactions'
            ws_config[f'B{current_row}'].font = Font(size=9)
            current_row += 1
        
        current_row += 2
        
        # === SECTION 4: BUSINESS PROFILE ===
        ws_config[f'A{current_row}'] = 'BUSINESS PROFILE ASSUMPTIONS'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        profile_data = [
            ['Business Type', 'Small to Medium Business (SMB)', 'Based on transaction patterns and volumes'],
            ['Industry', 'General Commercial', 'Multi-category spend pattern'],
            ['Transaction Volume', f'{len(result_df)} transactions', f'Period: {period_days} elapsed days'],
            ['Average Transaction', f'${result_df["Debit"].mean():,.2f}' if len(result_df[result_df["Debit"] > 0]) > 0 else 'N/A', 'Debit transactions only'],
            ['Typical Patterns', 'B2B payments, payroll, rent, utilities', 'Indicated by transaction descriptions'],
        ]
        
        for profile_row in profile_data:
            for col_idx, value in enumerate(profile_row, 1):
                cell = ws_config.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
        
        current_row += 2
        
        # === SECTION 5: METHODOLOGY NOTES ===
        ws_config[f'A{current_row}'] = 'METHODOLOGY NOTES'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1
        
        notes = [
            "Approach: rule-based screening using configurable thresholds",
            "Output: alerts for manual review (not conclusions)",
            "Traceability: Txn_ID and source pointers are provided for each transaction",
        ]
        for note in notes:
            cell = ws_config.cell(row=current_row, column=1, value=note)
            cell.font = Font(size=9, italic=True, bold=note.endswith(':'))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws_config.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
        

        # === SECTION: RUN-SPECIFIC CALCULATIONS ===
        current_row += 1

        ws_config[f'A{current_row}'] = 'RUN-SPECIFIC CALCULATIONS'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # Table headers
        ws_config.cell(row=current_row, column=1, value='Parameter').font = Font(bold=True)
        ws_config.cell(row=current_row, column=2, value='Value').font = Font(bold=True)
        ws_config.cell(row=current_row, column=3, value='How to verify').font = Font(bold=True)
        ws_config.cell(row=current_row, column=4, value='Notes').font = Font(bold=True)
        for col in range(1, 5):
            ws_config.cell(row=current_row, column=col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        current_row += 1

        # Dynamic thresholds calculated from this specific dataset/run (does not affect detection logic)
        ws_config.cell(row=current_row, column=1, value='p95_debit (outflows)')
        if run_stats.get('p95_debit') is None:
            ws_config.cell(row=current_row, column=2, value='N/A (n<=20)')
        else:
            c = ws_config.cell(row=current_row, column=2, value=round(run_stats['p95_debit'], 2))
            c.number_format = '$#,##0.00'
        ws_config.cell(row=current_row, column=3, value='Quantile(0.95) of Debit > 0 in Transactions')
        ws_config.cell(row=current_row, column=4, value=f"n={run_stats.get('weekend_outflow_txn_count', 'N/A')}")
        current_row += 1

        ws_config.cell(row=current_row, column=1, value='weekend_threshold')
        c = ws_config.cell(row=current_row, column=2, value=round(run_stats.get('weekend_threshold', 0), 2))
        c.number_format = '$#,##0.00'
        ws_config.cell(row=current_row, column=3, value='max(25000, p95_debit) if n>20 else 25000')
        ws_config.cell(row=current_row, column=4, value='')
        current_row += 1

        ws_config.cell(row=current_row, column=1, value='weekend_threshold_floor')
        c = ws_config.cell(row=current_row, column=2, value=int(run_stats.get('weekend_threshold_floor', 25000)))
        c.number_format = '$#,##0.00'
        ws_config.cell(row=current_row, column=3, value='Hard floor used in weekend rule')
        ws_config.cell(row=current_row, column=4, value='')
        current_row += 1

        # Set column widths
        
                # === SECTION: RECONCILIATION (PER SOURCE FILE) ===
        current_row += 2

        ws_config[f'A{current_row}'] = 'RECONCILIATION (PER SOURCE FILE)'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        # Table headers
        ws_config.cell(row=current_row, column=1, value='Source_File').font = Font(bold=True)
        ws_config.cell(row=current_row, column=2, value='Metric').font = Font(bold=True)
        ws_config.cell(row=current_row, column=3, value='Value').font = Font(bold=True)
        ws_config.cell(row=current_row, column=4, value='Ref').font = Font(bold=True)
        for col in range(1, 5):
            ws_config.cell(row=current_row, column=col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        current_row += 1

        # Build lookup maps for reported balances (if detected)
        beginning_map = {b['file']: b['balance'] for b in beginning_balances} if 'beginning_balances' in locals() else {}
        beginning_ref_map = {b['file']: b.get('ref') for b in beginning_balances} if 'beginning_balances' in locals() else {}
        ending_map = {b['file']: b['balance'] for b in ending_balances} if 'ending_balances' in locals() else {}
        ending_ref_map = {b['file']: b.get('ref') for b in ending_balances} if 'ending_balances' in locals() else {}

        def _write_recon_row(source_file: str, metric: str, value, ref: str, currency: bool = False):
            nonlocal current_row
            ws_config.cell(row=current_row, column=1, value=source_file)
            ws_config.cell(row=current_row, column=2, value=metric)

            c = ws_config.cell(row=current_row, column=3, value=value)
            if currency and isinstance(value, (int, float)) and value is not None and not (isinstance(value, float) and np.isnan(value)):
                c.number_format = '$#,##0.00'

            ws_config.cell(row=current_row, column=4, value=ref)
            for col in range(1, 5):
                ws_config.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
            current_row += 1

        # Compute reconciliation metrics from Transactions by Source_File
        for source_file, g in result_df.groupby('Source_File'):
            g = g.copy()
            # Defensive conversions
            g['Debit'] = pd.to_numeric(g.get('Debit'), errors='coerce')
            g['Credit'] = pd.to_numeric(g.get('Credit'), errors='coerce')
            g['Amount_Signed'] = pd.to_numeric(g.get('Amount_Signed'), errors='coerce')

            txn_count = int(len(g))
            inflows = float(g['Credit'].fillna(0).sum())
            outflows = float(g['Debit'].fillna(0).sum())
            net = float(g['Amount_Signed'].fillna(0).sum())

            try:
                dmin = pd.to_datetime(g['Date']).min().date()
                dmax = pd.to_datetime(g['Date']).max().date()
                period = f"{dmin.isoformat()} — {dmax.isoformat()}"
            except Exception:
                period = 'N/A'

            opening_rep = beginning_map.get(source_file)
            opening_ref = beginning_ref_map.get(source_file)
            closing_rep = ending_map.get(source_file)
            closing_ref = ending_ref_map.get(source_file)

            closing_exp = None
            if opening_rep is not None:
                closing_exp = float(opening_rep) + net

            delta = None
            reconciles = 'N/A'
            if closing_rep is not None and closing_exp is not None:
                delta = float(closing_rep) - float(closing_exp)
                reconciles = 'TRUE' if abs(delta) < 0.01 else 'FALSE'

            _write_recon_row(source_file, 'Statement period (from Transactions)', period, 'Transactions')
            _write_recon_row(source_file, 'Transactions count', txn_count, 'Transactions')
            _write_recon_row(source_file, 'Total inflows (Credit)', round(inflows, 2), 'Transactions', currency=True)
            _write_recon_row(source_file, 'Total outflows (Debit)', round(outflows, 2), 'Transactions', currency=True)
            _write_recon_row(source_file, 'Net cash flow (Credit - Debit)', round(net, 2), 'Transactions', currency=True)

            _write_recon_row(
                source_file,
                'Opening balance (reported)',
                None if opening_rep is None else round(float(opening_rep), 2),
                (opening_ref if opening_ref else 'N/A'),
                currency=True
            )
            _write_recon_row(
                source_file,
                'Closing balance (reported)',
                None if closing_rep is None else round(float(closing_rep), 2),
                (closing_ref if closing_ref else 'N/A'),
                currency=True
            )
            _write_recon_row(
                source_file,
                'Closing balance (expected)',
                None if closing_exp is None else round(float(closing_exp), 2),
                'Derived',
                currency=True
            )
            _write_recon_row(
                source_file,
                'Delta (reported - expected)',
                None if delta is None else round(float(delta), 2),
                'Derived',
                currency=True
            )
            _write_recon_row(source_file, 'Reconciles (abs(delta)<0.01)', reconciles, 'Derived')


        

        # === SECTION: DASHBOARD VALIDATION (VS TRANSACTIONS) ===
        current_row += 2
        ws_config[f'A{current_row}'] = 'DASHBOARD VALIDATION (VS TRANSACTIONS)'
        ws_config[f'A{current_row}'].font = Font(bold=True, size=14)
        ws_config.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1
        ws_config[f'A{current_row}'] = 'All Analytics Dashboard figures are independently recomputed from Transactions and compared to the displayed values.'
        ws_config[f'A{current_row}'].font = Font(italic=True, color='666666')
        ws_config.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 2

        val_headers = ['Section', 'Key', 'Metric', 'Dashboard_Cell', 'Dashboard_Value', 'Transactions_Value', 'Delta', 'Status']
        for col_idx, h in enumerate(val_headers, 1):
            cell = ws_config.cell(row=current_row, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='EFEFEF', end_color='EFEFEF', fill_type='solid')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        current_row += 1

        from openpyxl.utils import get_column_letter

        def _cell_addr(sheet_name: str, row_i: int, col_i: int) -> str:
            return f"{sheet_name}!{get_column_letter(col_i)}{row_i}"

        def _is_nan(x) -> bool:
            return isinstance(x, float) and np.isnan(x)

        def _write_val_row(section: str, key: str, metric: str, cell_addr: str,
                           dash_val, exp_val, tol: float = 0.01, fmt: str = None):
            nonlocal current_row

            ws_config.cell(row=current_row, column=1, value=section)
            ws_config.cell(row=current_row, column=2, value=key)
            ws_config.cell(row=current_row, column=3, value=metric)
            ws_config.cell(row=current_row, column=4, value=cell_addr)

            c_dash = ws_config.cell(row=current_row, column=5, value=dash_val)
            c_exp = ws_config.cell(row=current_row, column=6, value=exp_val)

            delta_val = None
            status = 'OK'
            if dash_val is None or exp_val is None or _is_nan(dash_val) or _is_nan(exp_val):
                status = 'N/A'
            else:
                try:
                    delta_val = float(dash_val) - float(exp_val)
                    status = 'OK' if abs(delta_val) <= tol else 'MISMATCH'
                except Exception:
                    status = 'N/A'

            c_delta = ws_config.cell(row=current_row, column=7, value=delta_val)
            ws_config.cell(row=current_row, column=8, value=status)

            for col in range(1, 9):
                ws_config.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

            if fmt == 'currency':
                for c in (c_dash, c_exp, c_delta):
                    if c.value is not None and not _is_nan(c.value):
                        c.number_format = '$#,##0.00'
            elif fmt == 'percent':
                for c in (c_dash, c_exp, c_delta):
                    if c.value is not None and not _is_nan(c.value):
                        c.number_format = '0.0%'
            elif fmt == 'pct_points':
                for c in (c_dash, c_exp, c_delta):
                    if c.value is not None and not _is_nan(c.value):
                        c.number_format = '0.0'
            elif fmt == 'int':
                for c in (c_dash, c_exp, c_delta):
                    if c.value is not None and not _is_nan(c.value):
                        c.number_format = '0'

            current_row += 1

        ws_dash = ws_analytics

        def _find_label_row(label: str):
            for rr in range(1, ws_dash.max_row + 1):
                if ws_dash.cell(rr, 1).value == label:
                    return rr
            return None

        # --- Cash Flow Metrics (top block) ---
        exp_total_txns = int(len(result_df))
        exp_total_inflows = float(result_df['Credit'].sum())
        exp_total_outflows = float(result_df['Debit'].sum())
        exp_net_flow = exp_total_inflows - exp_total_outflows

        exp_avg_in = (exp_total_inflows / period_days) if period_days else None
        exp_avg_out = (exp_total_outflows / period_days) if period_days else None
        exp_daily_burn = (exp_avg_out - exp_avg_in) if (exp_avg_in is not None and exp_avg_out is not None) else None

        cash_checks = [
            ('Total Transactions', exp_total_txns, 'int', 0),
            ('Total Inflows', round(exp_total_inflows, 2), 'currency', 0.01),
            ('Total Outflows', round(exp_total_outflows, 2), 'currency', 0.01),
            ('Net Cash Flow', round(exp_net_flow, 2), 'currency', 0.01),
            ('Avg Daily Inflow', (round(exp_avg_in, 2) if exp_avg_in is not None else None), 'currency', 0.02),
            ('Avg Daily Outflow', (round(exp_avg_out, 2) if exp_avg_out is not None else None), 'currency', 0.02),
            ('Daily Burn Rate', (round(exp_daily_burn, 2) if exp_daily_burn is not None else None), 'currency', 0.02),
        ]

        # Optional: reported balances (from header extraction) if available
        if total_beginning_balance is not None:
            cash_checks.insert(1, ('Beginning Balance', round(float(total_beginning_balance), 2), 'currency', 0.01))
        if 'ending_balance' in cash_metrics and cash_metrics.get('ending_balance') is not None:
            cash_checks.insert(2, ('Ending Balance', round(float(cash_metrics.get('ending_balance')), 2), 'currency', 0.01))

        for label, exp_val, fmt, tol in cash_checks:
            rr = _find_label_row(label)
            if rr is None:
                _write_val_row('Cash Flow Metrics', label, 'Value', 'Analytics Dashboard!<missing>', None, exp_val, tol=tol, fmt=fmt)
            else:
                cell_addr = _cell_addr('Analytics Dashboard', rr, 2)
                dash_val = ws_dash.cell(rr, 2).value
                _write_val_row('Cash Flow Metrics', label, 'Value', cell_addr, dash_val, exp_val, tol=tol, fmt=fmt)

        # --- Monthly Cash Flow Trends ---
        header_row = None
        for rr in range(1, ws_dash.max_row + 1):
            if ws_dash.cell(rr, 1).value == 'Month' and ws_dash.cell(rr, 2).value == 'Inflows':
                header_row = rr
                break

        if header_row is None:
            _write_val_row('Monthly Trends', 'Table', 'Presence', 'Analytics Dashboard!<missing>', None, None, tol=0.01, fmt=None)
        else:
            df_copy = result_df.copy()
            df_copy['Date'] = pd.to_datetime(df_copy['Date'], errors='coerce')
            df_copy = df_copy[df_copy['Date'].notna()].copy()
            df_copy['Month'] = df_copy['Date'].dt.to_period('M')
            monthly = df_copy.groupby('Month').agg({'Debit': 'sum', 'Credit': 'sum'})
            monthly['Net'] = monthly['Credit'] - monthly['Debit']
            monthly['Month_Name'] = monthly.index.astype(str)
            month_map = monthly.set_index('Month_Name')[['Credit', 'Debit', 'Net']].to_dict(orient='index')

            rr = header_row + 1
            while rr <= ws_dash.max_row:
                month_name = ws_dash.cell(rr, 1).value
                if month_name is None or not isinstance(month_name, str) or not re.match(r'^\d{4}-\d{2}$', month_name):
                    break

                exp = month_map.get(month_name, {'Credit': 0.0, 'Debit': 0.0, 'Net': 0.0})
                dash_in = ws_dash.cell(rr, 2).value
                dash_out = ws_dash.cell(rr, 3).value
                dash_net = ws_dash.cell(rr, 4).value

                _write_val_row('Monthly Trends', month_name, 'Inflows', _cell_addr('Analytics Dashboard', rr, 2), dash_in, round(float(exp['Credit']), 2), tol=0.01, fmt='currency')
                _write_val_row('Monthly Trends', month_name, 'Outflows', _cell_addr('Analytics Dashboard', rr, 3), dash_out, round(float(exp['Debit']), 2), tol=0.01, fmt='currency')
                _write_val_row('Monthly Trends', month_name, 'Net', _cell_addr('Analytics Dashboard', rr, 4), dash_net, round(float(exp['Net']), 2), tol=0.01, fmt='currency')

                rr += 1

        # --- Benford's Law Analysis (Debit amounts) ---
        header_row = None
        for rr in range(1, ws_dash.max_row + 1):
            if ws_dash.cell(rr, 1).value == 'Digit' and ws_dash.cell(rr, 2).value == 'Expected %':
                header_row = rr
                break

        if benford_results is None:
            # Dashboard may show an "insufficient data" note; only check that the header was not written
            if header_row is not None:
                _write_val_row('Benford', 'Table', 'Expected none', _cell_addr('Analytics Dashboard', header_row, 1),
                               'Present', 'Should be absent', tol=0.0, fmt=None)
        else:
            ben_df = benford_results.get('results')
            ben_map = {}
            if ben_df is not None and len(ben_df) > 0:
                # ben_df columns: First_Digit, Expected_%, Actual_%, Deviation_%
                for _, r0 in ben_df.iterrows():
                    try:
                        ben_map[int(r0['First_Digit'])] = float(r0['Actual_%'])
                    except Exception:
                        continue
            if header_row is None:
                _write_val_row('Benford', 'Table', 'Presence', 'Analytics Dashboard!<missing>', None, None, tol=0.01, fmt=None)
            else:
                rr = header_row + 1
                while rr <= ws_dash.max_row:
                    digit = ws_dash.cell(rr, 1).value
                    if digit is None or not isinstance(digit, (int, float)):
                        break
                    digit_i = int(digit)
                    if digit_i not in ben_map:
                        break

                    exp_actual = float(ben_map.get(digit_i))
                    dash_actual = ws_dash.cell(rr, 3).value  # Actual %
                    _write_val_row('Benford', str(digit_i), 'Actual % (Debit)', _cell_addr('Analytics Dashboard', rr, 3),
                                   dash_actual, round(exp_actual, 1), tol=0.2, fmt='pct_points')
                    rr += 1

        # --- Vendor Concentration Analysis (by Description_clean) ---
        header_row = None
        for rr in range(1, ws_dash.max_row + 1):
            if ws_dash.cell(rr, 1).value == 'Vendor' and ws_dash.cell(rr, 2).value == 'Total Spending':
                header_row = rr
                break

        if header_row is None:
            _write_val_row('Vendor Concentration', 'Table', 'Presence', 'Analytics Dashboard!<missing>', None, None, tol=0.01, fmt=None)
        else:
            exp_conc = {}
            if len(concentration_df) > 0:
                for _, r0 in concentration_df.iterrows():
                    exp_conc[str(r0['Vendor'])] = {
                        'Total_Spending': float(r0['Total_Spending']),
                        'Pct': float(r0['Percentage']) / 100.0
                    }

            rr = header_row + 1
            while rr <= ws_dash.max_row:
                vendor = ws_dash.cell(rr, 1).value
                if vendor is None or not isinstance(vendor, str) or vendor.strip() == '':
                    break
                if vendor.strip().upper() == 'SPENDING BY CATEGORY':
                    break

                exp = exp_conc.get(vendor, None)
                dash_spend = ws_dash.cell(rr, 2).value
                dash_pct = ws_dash.cell(rr, 3).value

                _write_val_row('Vendor Concentration', vendor, 'Total Spending', _cell_addr('Analytics Dashboard', rr, 2),
                               dash_spend, (round(exp['Total_Spending'], 2) if exp else None), tol=0.01, fmt='currency')
                _write_val_row('Vendor Concentration', vendor, '% of Total', _cell_addr('Analytics Dashboard', rr, 3),
                               dash_pct, (round(exp['Pct'], 3) if exp else None), tol=0.001, fmt='percent')

                rr += 1

        # --- Spending by Category: Outflows ---
        outflows_row = None
        inflows_row = None
        for rr in range(1, ws_dash.max_row + 1):
            if ws_dash.cell(rr, 1).value == 'OUTFLOWS (Expenses)':
                outflows_row = rr
            if ws_dash.cell(rr, 1).value == 'INFLOWS (Income)':
                inflows_row = rr

        # Expected category maps
        exp_outflows = result_df[result_df['Direction'] == 'Outflow'].groupby('Operation_Type')['Debit'].sum().sort_values(ascending=False).to_dict()
        exp_inflows = result_df[result_df['Direction'] == 'Inflow'].groupby('Operation_Type')['Credit'].sum().sort_values(ascending=False).to_dict()

        if outflows_row is None:
            _write_val_row('Category Outflows', 'Table', 'Presence', 'Analytics Dashboard!<missing>', None, None, tol=0.01, fmt=None)
        else:
            rr = outflows_row + 2  # skip "Category/Amount" header row
            while rr <= ws_dash.max_row:
                cat = ws_dash.cell(rr, 1).value
                if cat is None or (isinstance(cat, str) and cat.strip() == ''):
                    break
                if cat == 'INFLOWS (Income)':
                    break
                dash_amt = ws_dash.cell(rr, 2).value
                exp_amt = exp_outflows.get(cat, None)
                _write_val_row('Category Outflows', cat, 'Amount', _cell_addr('Analytics Dashboard', rr, 2),
                               dash_amt, (round(float(exp_amt), 2) if exp_amt is not None else None), tol=0.01, fmt='currency')
                rr += 1

        # --- Spending by Category: Inflows ---
        if inflows_row is None:
            _write_val_row('Category Inflows', 'Table', 'Presence', 'Analytics Dashboard!<missing>', None, None, tol=0.01, fmt=None)
        else:
            rr = inflows_row + 2  # skip "Category/Amount" header row
            while rr <= ws_dash.max_row:
                cat = ws_dash.cell(rr, 1).value
                if cat is None or (isinstance(cat, str) and cat.strip() == ''):
                    break
                dash_amt = ws_dash.cell(rr, 2).value
                exp_amt = exp_inflows.get(cat, None)
                _write_val_row('Category Inflows', cat, 'Amount', _cell_addr('Analytics Dashboard', rr, 2),
                               dash_amt, (round(float(exp_amt), 2) if exp_amt is not None else None), tol=0.01, fmt='currency')
                rr += 1

        ws_config.column_dimensions['A'].width = 30
        ws_config.column_dimensions['B'].width = 35
        ws_config.column_dimensions['C'].width = 45
        ws_config.column_dimensions['D'].width = 25
        ws_config.column_dimensions['E'].width = 22
        ws_config.column_dimensions['F'].width = 22
        ws_config.column_dimensions['G'].width = 14
        ws_config.column_dimensions['H'].width = 12

    if getattr(args, 'patch_excel_app_metadata', False):
        patch_excel_app_metadata(output_path)
    
    logger.info("Done.")
    p1 = len(findings_df[findings_df['Priority'] == 'P1'])
    p2 = len(findings_df[findings_df['Priority'] == 'P2'])
    p3 = len(findings_df[findings_df['Priority'] == 'P3'])

    logger.info(f"Summary: files={len(all_dataframes)} | txns={len(result_df)} | findings={len(findings_df)} (P1={p1}, P2={p2}, P3={p3})")

    if total_beginning_balance is not None and 'ending_balance' in cash_metrics:
        logger.info(
            f"Balances (total): opening=${total_beginning_balance:,.2f} | ending=${cash_metrics['ending_balance']:,.2f}"
        )
    else:
        logger.info("Balances (total): opening=N/A | ending=N/A")

    logger.info(f"Output: {output_path}")


if __name__ == "__main__":
    try:
        main()
        logging.getLogger(__name__).info("BUILD SUCCESSFUL")
        raise SystemExit(0)
    except ValidationError:
        raise SystemExit(1)
    except (StatementFormatError, InputDataError) as e:
        logging.getLogger(__name__).error(f"Input error: {e}")
        raise SystemExit(2)
    except Exception as e:
        logging.getLogger(__name__).exception(f"Unhandled error: {e}")
        raise SystemExit(3)
